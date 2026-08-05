import asyncio
import io
import logging
import math
import os
import re
import sqlite3
import zipfile
from contextlib import contextmanager
from datetime import date, datetime, timedelta
from pathlib import Path
from zoneinfo import ZoneInfo, ZoneInfoNotFoundError

import aiohttp
import openpyxl
import xlrd
from aiogram import BaseMiddleware, Bot, Dispatcher, F
from aiogram.filters import Command
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.fsm.storage.memory import MemoryStorage
from aiogram.types import (
    Document,
    KeyboardButton,
    Message,
    ReplyKeyboardMarkup,
)
from dotenv import load_dotenv

from balance_logic import (
    FIELDS,
    bal1,
    bal2,
    balance_status,
    balc1,
    balc2,
    format_period_label,
    infer_report_period,
    is_consecutive_period,
    rolling_snapshots,
    sum_period,
)

BASE_DIR = Path(__file__).resolve().parent
load_dotenv(BASE_DIR / ".env")
logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
logger = logging.getLogger(__name__)

# ════════════════════════════════════════════════════════
#  НАСТРОЙКИ
# ════════════════════════════════════════════════════════
BOT_TOKEN = os.getenv("BOT_TOKEN")
ANTHROPIC_API_KEY = os.getenv("ANTHROPIC_API_KEY")
AI_MODEL = os.getenv("AI_MODEL", "claude-sonnet-4-6")

db_path_setting = Path(os.getenv("DB_PATH", "dof_balance.db"))
DB_PATH = str(
    db_path_setting if db_path_setting.is_absolute() else BASE_DIR / db_path_setting
)

BOT_TIMEZONE_NAME = os.getenv("BOT_TIMEZONE", "Asia/Qostanay")
try:
    BOT_TIMEZONE = ZoneInfo(BOT_TIMEZONE_NAME)
except ZoneInfoNotFoundError as exc:
    raise RuntimeError(
        f"Неизвестный часовой пояс BOT_TIMEZONE={BOT_TIMEZONE_NAME!r}"
    ) from exc

try:
    MAX_REPORT_SIZE_MB = int(os.getenv("MAX_REPORT_SIZE_MB", "15"))
except ValueError as exc:
    raise RuntimeError("MAX_REPORT_SIZE_MB должен быть целым числом") from exc
if MAX_REPORT_SIZE_MB <= 0:
    raise RuntimeError("MAX_REPORT_SIZE_MB должен быть больше нуля")
MAX_REPORT_BYTES = MAX_REPORT_SIZE_MB * 1024 * 1024

try:
    MAX_XLSX_UNCOMPRESSED_MB = int(os.getenv("MAX_XLSX_UNCOMPRESSED_MB", "100"))
except ValueError as exc:
    raise RuntimeError("MAX_XLSX_UNCOMPRESSED_MB должен быть целым числом") from exc
if MAX_XLSX_UNCOMPRESSED_MB <= 0:
    raise RuntimeError("MAX_XLSX_UNCOMPRESSED_MB должен быть больше нуля")
MAX_XLSX_UNCOMPRESSED_BYTES = MAX_XLSX_UNCOMPRESSED_MB * 1024 * 1024
MAX_XLSX_ARCHIVE_ENTRIES = 10_000


def _parse_allowed_user_ids(raw: str) -> set[int]:
    if not raw.strip():
        return set()
    try:
        return {int(item) for item in re.split(r"[\s,;]+", raw.strip()) if item}
    except ValueError as exc:
        raise RuntimeError(
            "ALLOWED_USER_IDS должен содержать Telegram ID через запятую"
        ) from exc


ALLOWED_USER_IDS = _parse_allowed_user_ids(os.getenv("ALLOWED_USER_IDS", ""))

if not BOT_TOKEN:
    raise RuntimeError("BOT_TOKEN не задан (переменная окружения)")


def local_now() -> datetime:
    return datetime.now(BOT_TIMEZONE)


# ════════════════════════════════════════════════════════
#  НОРМЫ БАЛАНСА (% от Конв.3 или Конв.4 = 100%)
#  ключ: (мин%, макс%, человекочитаемое имя, очередь)
# ════════════════════════════════════════════════════════
NORMS = {
    "kv14": (60, 80, "Конв.14 (дробление/грохочение)", 1),
    "kv15": (60, 80, "Конв.15 (дробление/грохочение)", 2),
    "kv32": (40, 60, "Конв.32 (ПП после сепарации)", 1),
    "kv31": (40, 60, "Конв.31 (ПП после сепарации)", 2),
    "kv34": (83, 87, "Конв.34 (ПП → ММО)", 1),
    "kv33": (83, 87, "Конв.33 (ПП → ММО)", 2),
    "kv102": (6, 20, "Конв.102 (хвосты → склад 104)", 1),
    "kv101": (6, 20, "Конв.101 (хвосты → склад 104)", 2),
    "kv19": (40, 60, "Конв.19 (ПП 2 очереди)", 2),
    "kv34a": (15, 45, "Конв.10/34А (подрешетный продукт)", 1),
}
WARN_PCT = 2.0  # баланс: предупреждение
CRIT_PCT = 5.0  # баланс: критично
DUP_WARN = 0.5  # дублирующие весы: предупреждение, %
DUP_CRIT = 2.0  # дублирующие весы: критично, %

# Маппинг названий конвейеров из отчёта → ключи БД
CONV_MAP = {
    "Конвейер 4": "kv4",
    "Конвейер 4Д": "kv4d",
    "Конвейер 14": "kv14",
    "Конвейер 32": "kv32",
    "Конвейер 18": "kv18",
    "Конвейер 102": "kv102",
    "Конвейер 34": "kv34",
    "Конвейер 24ПП": "kv24p",
    "Конвейер 24ХВ": "kv24hv",
    "Конвейер 28А.1": "kv28a1",
    "Конвейер 34A": "kv34a",
    "Конвейер 34А": "kv34a",
    "Конвейер 3": "kv3",
    "Конвейер 3Д": "kv3d",
    "Конвейер 15": "kv15",
    "Конвейер 31": "kv31",
    "Конвейер 19": "kv19",
    "Конвейер 101": "kv101",
    "Конвейер 33": "kv33",
    "Конвейер 28А.2": "kv28a2",
    "Конвейер 44": "kv44",
    "Конвейер 44Д": "kv44d",
    "Конвейер 46": "kv46",
    "Конвейер 46Д": "kv46d",
    "Конвейер 74": "kv74",
    "Конвейер 74Д": "kv74d",
    "Конвейер 65МПС": "kv65mps",
    "Конвейер 65ЦПО": "kv65cpo",
    "Конвейер 66МПС": "kv66mps",
    "Конвейер 66ЦПО": "kv66cpo",
    "Конвейер 84МПС": "kv84mps",
    "Конвейер 84ЦПО": "kv84cpo",
    "Конвейер 63": "kv63",
    "Конвейер 61": "kv61",
}

# Допустимые варианты обозначений, встречающиеся в выгрузках Excel.
CONV_MAP.update(
    {
        "Конвейер 28А I": "kv28a1",
        "Конвейер 28А I очередь": "kv28a1",
        "Конвейер 28А II": "kv28a2",
        "Конвейер 28А II очередь": "kv28a2",
    }
)


def _normalize_label(value) -> str:
    text = str(value or "").replace("\xa0", " ").strip().casefold()
    text = text.replace("ё", "е")
    return re.sub(r"[\s._:()\-/]+", "", text)


NORMALIZED_CONV_MAP = {_normalize_label(name): key for name, key in CONV_MAP.items()}

REQUIRED_BASE_FIELDS = {"kv4", "kv3"}
BALANCE_REPORT_FIELDS = {
    "kv4",
    "kv14",
    "kv32",
    "kv34",
    "kv102",
    "kv24p",
    "kv24hv",
    "kv28a1",
    "kv3",
    "kv15",
    "kv31",
    "kv33",
    "kv101",
    "kv28a2",
}
NORM_REPORT_FIELDS = set(NORMS)
DUPLICATE_REPORT_FIELDS = {"kv4d", "kv3d"}
STOCK_REPORT_FIELDS = {
    "kv44",
    "kv44d",
    "kv46d",
    "kv74",
    "kv74d",
    "kv65mps",
    "kv65cpo",
    "kv66mps",
    "kv66cpo",
    "kv84mps",
    "kv84cpo",
}
FIELD_LABELS = {
    "kv4": "Конвейер 4",
    "kv4d": "Конвейер 4Д",
    "kv14": "Конвейер 14",
    "kv32": "Конвейер 32",
    "kv34": "Конвейер 34",
    "kv34a": "Конвейер 10/34А",
    "kv102": "Конвейер 102",
    "kv24p": "Конвейер 24ПП",
    "kv24hv": "Конвейер 24ХВ",
    "kv28a1": "Конвейер 28А.I",
    "kv3": "Конвейер 3",
    "kv3d": "Конвейер 3Д",
    "kv15": "Конвейер 15",
    "kv19": "Конвейер 19",
    "kv31": "Конвейер 31",
    "kv33": "Конвейер 33",
    "kv101": "Конвейер 101",
    "kv28a2": "Конвейер 28А.II",
    "kv44": "Конвейер 44",
    "kv44d": "Конвейер 44Д",
    "kv46d": "Конвейер 46Д",
    "kv74": "Конвейер 74",
    "kv74d": "Конвейер 74Д",
    "kv65mps": "Конвейер 65МПС",
    "kv65cpo": "Конвейер 65ЦПО",
    "kv66mps": "Конвейер 66МПС",
    "kv66cpo": "Конвейер 66ЦПО",
    "kv84mps": "Конвейер 84МПС",
    "kv84cpo": "Конвейер 84ЦПО",
}

# ════════════════════════════════════════════════════════
#  ФОРМУЛЫ СКЛАДА ВЛАЖНОГО КОНЦЕНТРАТА
# ════════════════════════════════════════════════════════
STOCK_DIFF_WARN = 500  # тонн — порог алерта по расхождению
ACTIVE_REPORT_SCOPE_ID = 0  # одна общая производственная база на экземпляр бота


def finite_number(value, default: float = 0.0) -> float:
    try:
        number = float(value)
    except (TypeError, ValueError):
        return default
    return number if math.isfinite(number) else default


def calc_produced(d: dict) -> float:
    """Произведено = (44+44Д)/2 + 46Д + (74+74Д)/2"""
    avg44 = (finite_number(d.get("kv44")) + finite_number(d.get("kv44d"))) / 2
    k46d = finite_number(d.get("kv46d"))
    avg74 = (finite_number(d.get("kv74")) + finite_number(d.get("kv74d"))) / 2
    return avg44 + k46d + avg74


def calc_shipped(d: dict) -> float:
    """Отгружено = 65МПС+65ЦПО+66МПС+66ЦПО+84МПС+84ЦПО"""
    keys = ["kv65mps", "kv65cpo", "kv66mps", "kv66cpo", "kv84mps", "kv84cpo"]
    return sum(finite_number(d.get(key)) for key in keys)


# ════════════════════════════════════════════════════════
#  БАЗА ДАННЫХ (безопасные context manager'ы)
# ════════════════════════════════════════════════════════


@contextmanager
def db_connection():
    """SQLite-соединение с ожиданием блокировки и гарантированным закрытием."""
    db_file = Path(DB_PATH)
    if DB_PATH != ":memory:":
        db_file.parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(DB_PATH, timeout=30)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA busy_timeout=30000")
    conn.execute("PRAGMA foreign_keys=ON")
    try:
        yield conn
    except Exception:
        conn.rollback()
        raise
    else:
        conn.commit()
    finally:
        conn.close()


def _ensure_columns(conn: sqlite3.Connection, table: str, columns: dict[str, str]):
    existing = {
        row["name"] for row in conn.execute(f'PRAGMA table_info("{table}")').fetchall()
    }
    for column, definition in columns.items():
        if column not in existing:
            conn.execute(f'ALTER TABLE "{table}" ADD COLUMN "{column}" {definition}')


def init_db():
    with db_connection() as conn:
        conn.execute("PRAGMA journal_mode=WAL")
        conn.execute("PRAGMA synchronous=NORMAL")
        cols_sql = ", ".join(f"{f} REAL DEFAULT 0" for f in FIELDS)
        conn.execute(f"""
            CREATE TABLE IF NOT EXISTS daily_data (
                id          INTEGER PRIMARY KEY AUTOINCREMENT,
                report_date TEXT,
                day_num     INTEGER,
                year        INTEGER,
                month       INTEGER,
                {cols_sql},
                source      TEXT,
                uploaded    TEXT DEFAULT CURRENT_TIMESTAMP,
                UNIQUE(year, month, day_num)
            )
        """)
        conn.execute(f"""
            CREATE TABLE IF NOT EXISTS night_shift (
                id          INTEGER PRIMARY KEY AUTOINCREMENT,
                year        INTEGER,
                month       INTEGER,
                day_num     INTEGER,
                {cols_sql},
                source      TEXT,
                uploaded    TEXT DEFAULT CURRENT_TIMESTAMP,
                UNIQUE(year, month, day_num)
            )
        """)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS report_log (
                id         INTEGER PRIMARY KEY AUTOINCREMENT,
                filename   TEXT, period TEXT, rows_saved INTEGER,
                user_id    INTEGER, uploaded TEXT DEFAULT CURRENT_TIMESTAMP
            )
        """)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS stock_data (
                id              INTEGER PRIMARY KEY AUTOINCREMENT,
                year            INTEGER,
                month           INTEGER,
                day_num         INTEGER,
                stock_prev      REAL,
                stock_curr      REAL,
                produced        REAL,
                shipped         REAL,
                ves_izm         REAL,
                marksh_izm      REAL,
                nesovpadenie    REAL,
                user_id         INTEGER,
                entered_at      TEXT DEFAULT CURRENT_TIMESTAMP,
                UNIQUE(year, month, day_num)
            )
        """)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS active_report (
                user_id             INTEGER PRIMARY KEY,
                uploaded_by         INTEGER,
                year                INTEGER NOT NULL,
                month               INTEGER NOT NULL,
                last_completed_day  INTEGER,
                last_seen_day       INTEGER,
                filename            TEXT,
                period              TEXT,
                updated             TEXT DEFAULT CURRENT_TIMESTAMP
            )
        """)

        # Миграция старых рабочих баз: CREATE TABLE IF NOT EXISTS не добавляет
        # новые столбцы в уже существующую таблицу.
        common_period_columns = {
            "report_date": "TEXT",
            "day_num": "INTEGER",
            "year": "INTEGER",
            "month": "INTEGER",
            "source": "TEXT",
            "uploaded": "TEXT",
            **{field: "REAL DEFAULT 0" for field in FIELDS},
        }
        _ensure_columns(conn, "daily_data", common_period_columns)
        _ensure_columns(
            conn,
            "night_shift",
            {
                key: value
                for key, value in common_period_columns.items()
                if key != "report_date"
            },
        )
        _ensure_columns(
            conn,
            "report_log",
            {
                "filename": "TEXT",
                "period": "TEXT",
                "rows_saved": "INTEGER",
                "user_id": "INTEGER",
                "uploaded": "TEXT",
            },
        )
        _ensure_columns(
            conn,
            "stock_data",
            {
                "year": "INTEGER",
                "month": "INTEGER",
                "day_num": "INTEGER",
                "stock_prev": "REAL",
                "stock_curr": "REAL",
                "produced": "REAL",
                "shipped": "REAL",
                "ves_izm": "REAL",
                "marksh_izm": "REAL",
                "nesovpadenie": "REAL",
                "user_id": "INTEGER",
                "entered_at": "TEXT",
            },
        )
        _ensure_columns(
            conn,
            "active_report",
            {
                "uploaded_by": "INTEGER",
                "year": "INTEGER",
                "month": "INTEGER",
                "last_completed_day": "INTEGER",
                "last_seen_day": "INTEGER",
                "filename": "TEXT",
                "period": "TEXT",
                "updated": "TEXT",
            },
        )
        conn.execute(
            "CREATE INDEX IF NOT EXISTS idx_daily_report_date ON daily_data(report_date)"
        )
        conn.execute(
            """
            CREATE UNIQUE INDEX IF NOT EXISTS ux_daily_period_day
            ON daily_data(year, month, day_num)
            """
        )
        conn.execute(
            """
            CREATE UNIQUE INDEX IF NOT EXISTS ux_night_period_day
            ON night_shift(year, month, day_num)
            """
        )
        conn.execute(
            """
            CREATE UNIQUE INDEX IF NOT EXISTS ux_stock_period_day
            ON stock_data(year, month, day_num)
            """
        )


class ReportDataError(ValueError):
    """Отчёт прочитан, но его данные нельзя безопасно сохранить."""


def _shift_has_measurements(values: dict) -> bool:
    return any(finite_number(value) != 0 for value in values.values())


def _day_has_measurements(shifts: dict) -> bool:
    return _shift_has_measurements(shifts.get(1, {})) or _shift_has_measurements(
        shifts.get(2, {})
    )


def db_save_daily(
    parsed: dict,
    user_id: int,
    filename: str,
    year: int,
    month: int,
    today: date | None = None,
) -> tuple[int, int, int | None]:
    """Атомарно сохраняет завершённые сутки и текущую первую смену.

    Пустые колонки текущего и будущих дней игнорируются. Завершённые нулевые
    сутки сохраняются как возможная остановка. Ненулевые будущие даты
    отклоняются, а текущая дата не попадает в суточный баланс.
    """
    today = today or local_now().date()
    by_day = parsed.get("daily_by_shift", {})

    valid_days: list[tuple[int, date]] = []
    measured_days: list[tuple[int, date]] = []
    all_valid_by_day: dict[int, dict] = {}
    for raw_day, shifts in by_day.items():
        has_measurements = _day_has_measurements(shifts)
        try:
            day_num = int(raw_day)
            report_day = date(year, month, day_num)
        except (TypeError, ValueError) as exc:
            # Шаблоны февраля могут содержать полностью пустые колонки 29–31.
            if not has_measurements:
                continue
            raise ReportDataError(
                f"Некорректная дата в отчёте: {raw_day}.{month:02d}.{year}"
            ) from exc
        if report_day > today:
            if not has_measurements:
                continue
            raise ReportDataError(
                f"В отчёте найдены ненулевые данные за будущую дату "
                f"{report_day:%d.%m.%Y}. Проверьте период файла."
            )
        all_valid_by_day[day_num] = shifts
        valid_days.append((day_num, report_day))
        if has_measurements:
            measured_days.append((day_num, report_day))

    if not measured_days:
        raise ReportDataError(
            "В отчёте не найдено ни одного дня с ненулевыми показаниями весов."
        )

    # Все прошедшие нулевые сутки сохраняются как возможная остановка.
    # Пустые колонки текущего и будущих дней не считаются фактическими данными.
    data_days = [
        (day_num, report_day)
        for day_num, report_day in valid_days
        if report_day < today or _day_has_measurements(all_valid_by_day[day_num])
    ]
    data_days.sort()
    by_day = {day_num: all_valid_by_day[day_num] for day_num, _report_day in data_days}
    max_day = data_days[-1][0]
    incomplete_day = next(
        (day_num for day_num, report_day in data_days if report_day == today),
        None,
    )
    completed_days = [
        day_num for day_num, report_day in data_days if report_day < today
    ]

    base_fields = ["year", "month", "day_num", "report_date", "source"] + FIELDS
    cols = ",".join(base_fields)
    qs = ",".join(["?"] * len(base_fields))
    update_fields = (
        ",".join(f"{field}=excluded.{field}" for field in FIELDS)
        + ",report_date=excluded.report_date,source=excluded.source,"
        "uploaded=CURRENT_TIMESTAMP"
    )
    # Имена столбцов берутся только из константы FIELDS; значения параметризованы.
    sql_query = (
        f"INSERT INTO daily_data ({cols}) VALUES ({qs}) "  # nosec B608
        f"ON CONFLICT(year,month,day_num) DO UPDATE SET {update_fields}"
    )

    night_fields = ["year", "month", "day_num", "source"] + FIELDS
    night_cols = ",".join(night_fields)
    night_qs = ",".join(["?"] * len(night_fields))
    night_update = (
        ",".join(f"{field}=excluded.{field}" for field in FIELDS)
        + ",source=excluded.source,uploaded=CURRENT_TIMESTAMP"
    )
    night_sql = (
        f"INSERT INTO night_shift ({night_cols}) VALUES ({night_qs}) "  # nosec B608
        f"ON CONFLICT(year,month,day_num) DO UPDATE SET {night_update}"
    )

    with db_connection() as conn:
        conn.execute("BEGIN IMMEDIATE")

        # Лист «текущий месяц» является снимком всего периода. Замена месяца
        # целиком предотвращает возврат строк из более старого файла.
        conn.execute("DELETE FROM daily_data WHERE year=? AND month=?", (year, month))
        conn.execute("DELETE FROM night_shift WHERE year=? AND month=?", (year, month))

        for day_num in completed_days:
            shifts = by_day[day_num]
            shift_1 = shifts.get(1, {})
            shift_2 = shifts.get(2, {})
            values = {
                field: finite_number(shift_1.get(field))
                + finite_number(shift_2.get(field))
                for field in FIELDS
            }
            record = {
                "year": year,
                "month": month,
                "day_num": day_num,
                "report_date": f"{year:04d}-{month:02d}-{day_num:02d}",
                "source": filename,
                **values,
            }
            conn.execute(sql_query, list(record.values()))

        if incomplete_day is not None:
            shift_1 = by_day[incomplete_day].get(1, {})
            if _shift_has_measurements(shift_1):
                night_record = {
                    "year": year,
                    "month": month,
                    "day_num": incomplete_day,
                    "source": filename,
                    **{field: finite_number(shift_1.get(field)) for field in FIELDS},
                }
                conn.execute(night_sql, list(night_record.values()))

        # Пользовательские остатки склада сохраняются, но все производные
        # показатели пересчитываются по свежим весовым данным. Иначе после
        # исправленной загрузки раздел склада продолжал показывать старый расчёт.
        stock_rows = conn.execute(
            """
            SELECT id, day_num, stock_prev, stock_curr
            FROM stock_data
            WHERE year=? AND month=?
            """,
            (year, month),
        ).fetchall()
        for stock_row in stock_rows:
            daily_row = conn.execute(
                """
                SELECT * FROM daily_data
                WHERE year=? AND month=? AND day_num=?
                """,
                (year, month, stock_row["day_num"]),
            ).fetchone()
            if daily_row is None:
                continue
            fresh_data = dict(daily_row)
            produced = calc_produced(fresh_data)
            shipped = calc_shipped(fresh_data)
            ves_izm = produced - shipped
            marksh_izm = finite_number(stock_row["stock_curr"]) - finite_number(
                stock_row["stock_prev"]
            )
            conn.execute(
                """
                UPDATE stock_data
                SET produced=?, shipped=?, ves_izm=?, marksh_izm=?,
                    nesovpadenie=?, entered_at=CURRENT_TIMESTAMP
                WHERE id=?
                """,
                (
                    produced,
                    shipped,
                    ves_izm,
                    marksh_izm,
                    ves_izm - marksh_izm,
                    stock_row["id"],
                ),
            )

        last_completed_day = max(completed_days) if completed_days else None
        conn.execute(
            "INSERT INTO report_log (filename,period,rows_saved,user_id) VALUES (?,?,?,?)",
            (filename, parsed.get("period", ""), len(completed_days), user_id),
        )
        conn.execute(
            """
            INSERT INTO active_report
                (user_id, uploaded_by, year, month, last_completed_day,
                 last_seen_day, filename, period, updated)
            VALUES (?,?,?,?,?,?,?,?,CURRENT_TIMESTAMP)
            ON CONFLICT(user_id) DO UPDATE SET
                uploaded_by=excluded.uploaded_by,
                year=excluded.year,
                month=excluded.month,
                last_completed_day=excluded.last_completed_day,
                last_seen_day=excluded.last_seen_day,
                filename=excluded.filename,
                period=excluded.period,
                updated=CURRENT_TIMESTAMP
            """,
            (
                ACTIVE_REPORT_SCOPE_ID,
                user_id,
                year,
                month,
                last_completed_day,
                max_day,
                filename,
                parsed.get("period", ""),
            ),
        )

    return len(completed_days), max_day, incomplete_day


def db_get_active_report(_user_id: int) -> dict | None:
    """Возвращает общий последний успешно разобранный производственный файл."""
    with db_connection() as conn:
        row = conn.execute(
            "SELECT * FROM active_report WHERE user_id=?",
            (ACTIVE_REPORT_SCOPE_ID,),
        ).fetchone()
        # Совместимость с базой предыдущей версии до первой свежей загрузки.
        if row is None:
            row = conn.execute(
                "SELECT * FROM active_report ORDER BY updated DESC LIMIT 1"
            ).fetchone()
    return dict(row) if row else None


def active_report_end_date(active: dict | None) -> date | None:
    """Дата последнего завершённого дня активного отчёта."""
    if not active or active.get("last_completed_day") is None:
        return None
    try:
        return date(
            int(active["year"]),
            int(active["month"]),
            int(active["last_completed_day"]),
        )
    except (KeyError, TypeError, ValueError):
        return None


def db_get_night_shift(year: int, month: int) -> dict | None:
    with db_connection() as conn:
        row = conn.execute(
            "SELECT * FROM night_shift WHERE year=? AND month=? ORDER BY day_num DESC LIMIT 1",
            (year, month),
        ).fetchone()
    return dict(row) if row else None


def db_save_stock(
    year: int,
    month: int,
    day_num: int,
    stock_prev: float,
    stock_curr: float,
    user_id: int,
) -> dict:
    if not math.isfinite(stock_prev) or not math.isfinite(stock_curr):
        raise ValueError("Запасы должны быть конечными числами")

    with db_connection() as conn:
        conn.execute("BEGIN IMMEDIATE")
        row = conn.execute(
            "SELECT * FROM daily_data WHERE year=? AND month=? AND day_num=?",
            (year, month, day_num),
        ).fetchone()
        if not row:
            raise ReportDataError(
                f"Нет весовых данных за {day_num:02d}.{month:02d}.{year}"
            )

        data = dict(row)
        produced = calc_produced(data)
        shipped = calc_shipped(data)
        ves_izm = produced - shipped
        marksh_izm = stock_curr - stock_prev
        nesovpadenie = ves_izm - marksh_izm

        conn.execute(
            """
            INSERT INTO stock_data
                (year, month, day_num, stock_prev, stock_curr,
                 produced, shipped, ves_izm, marksh_izm, nesovpadenie, user_id)
            VALUES (?,?,?,?,?,?,?,?,?,?,?)
            ON CONFLICT(year,month,day_num) DO UPDATE SET
                stock_prev=excluded.stock_prev,
                stock_curr=excluded.stock_curr,
                produced=excluded.produced,
                shipped=excluded.shipped,
                ves_izm=excluded.ves_izm,
                marksh_izm=excluded.marksh_izm,
                nesovpadenie=excluded.nesovpadenie,
                user_id=excluded.user_id,
                entered_at=CURRENT_TIMESTAMP
        """,
            (
                year,
                month,
                day_num,
                stock_prev,
                stock_curr,
                produced,
                shipped,
                ves_izm,
                marksh_izm,
                nesovpadenie,
                user_id,
            ),
        )

    return {
        "day_num": day_num,
        "produced": produced,
        "shipped": shipped,
        "ves_izm": ves_izm,
        "marksh_izm": marksh_izm,
        "nesovpadenie": nesovpadenie,
    }


def db_get_stock_history(
    year: int,
    month: int,
    up_to_day: int | None = None,
) -> list:
    with db_connection() as conn:
        if up_to_day is None:
            rows = conn.execute(
                "SELECT * FROM stock_data WHERE year=? AND month=? ORDER BY day_num ASC",
                (year, month),
            ).fetchall()
        else:
            rows = conn.execute(
                """
                SELECT * FROM stock_data
                WHERE year=? AND month=? AND day_num<=?
                ORDER BY day_num ASC
                """,
                (year, month, up_to_day),
            ).fetchall()
    return [dict(r) for r in rows]


def db_get_month_data(
    year: int,
    month: int,
    up_to_day: int | None = None,
) -> list:
    with db_connection() as conn:
        if up_to_day is None:
            rows = conn.execute(
                "SELECT * FROM daily_data WHERE year=? AND month=? ORDER BY day_num ASC",
                (year, month),
            ).fetchall()
        else:
            rows = conn.execute(
                """
                SELECT * FROM daily_data
                WHERE year=? AND month=? AND day_num<=?
                ORDER BY day_num ASC
                """,
                (year, month, up_to_day),
            ).fetchall()
    return [dict(r) for r in rows]


def db_get_active_month_data(active: dict | None) -> list:
    """Данные только до последнего завершённого дня свежей загрузки."""
    if not active or active.get("last_completed_day") is None:
        return []
    return db_get_month_data(
        active["year"], active["month"], int(active["last_completed_day"])
    )


def db_get_active_stock_history(active: dict | None) -> list:
    if not active or active.get("last_completed_day") is None:
        return []
    with db_connection() as conn:
        rows = conn.execute(
            """
            SELECT stock_data.*
            FROM stock_data
            INNER JOIN daily_data USING (year, month, day_num)
            WHERE stock_data.year=? AND stock_data.month=?
              AND stock_data.day_num<=?
            ORDER BY stock_data.day_num ASC
            """,
            (
                active["year"],
                active["month"],
                int(active["last_completed_day"]),
            ),
        ).fetchall()
    return [dict(row) for row in rows]


def db_get_latest_completed_days(
    limit: int = 3,
    end_date: date | None = None,
) -> list:
    """Последние завершённые сутки не позднее активного отчёта."""
    limit = max(1, int(limit))
    with db_connection() as conn:
        if end_date is None:
            rows = conn.execute(
                """
                SELECT * FROM daily_data
                ORDER BY COALESCE(
                    NULLIF(report_date, ''),
                    printf('%04d-%02d-%02d', year, month, day_num)
                ) DESC
                LIMIT ?
                """,
                (limit,),
            ).fetchall()
        else:
            rows = conn.execute(
                """
                SELECT * FROM daily_data
                WHERE COALESCE(
                    NULLIF(report_date, ''),
                    printf('%04d-%02d-%02d', year, month, day_num)
                ) <= ?
                ORDER BY COALESCE(
                    NULLIF(report_date, ''),
                    printf('%04d-%02d-%02d', year, month, day_num)
                ) DESC
                LIMIT ?
                """,
                (end_date.isoformat(), limit),
            ).fetchall()
    return [dict(row) for row in reversed(rows)]


# ════════════════════════════════════════════════════════
#  ПАРСЕР EXCEL
# ════════════════════════════════════════════════════════
ZERO_CELL_MARKERS = {"", "-", "—", "–"}


def _parse_cell_number(value) -> float:
    if isinstance(value, UncachedFormula):
        raise TypeError(
            "формула не имеет сохранённого числового результата; "
            "откройте файл в Excel, выполните пересчёт и сохраните его заново"
        )
    if value is None:
        return 0.0
    if isinstance(value, str):
        normalized = value.replace("\xa0", " ").strip()
        if normalized in ZERO_CELL_MARKERS:
            return 0.0
        normalized = normalized.replace(" ", "").replace(",", ".")
    else:
        normalized = value
    try:
        number = float(normalized)
    except (TypeError, ValueError) as exc:
        raise ValueError(f"нечисловое значение {value!r}") from exc
    if not math.isfinite(number):
        raise ValueError(f"неконечное значение {value!r}")
    return number


def _read_xls_sheet(file_bytes: bytes, sheet_name: str) -> list:
    wb = xlrd.open_workbook(file_contents=file_bytes)
    try:
        if sheet_name not in wb.sheet_names():
            return []
        ws = wb.sheet_by_name(sheet_name)
        rows = []
        for row_index in range(ws.nrows):
            row = []
            for column_index in range(ws.ncols):
                cell = ws.cell(row_index, column_index)
                if cell.ctype == xlrd.XL_CELL_DATE:
                    row.append(xlrd.xldate_as_datetime(cell.value, wb.datemode))
                else:
                    row.append(cell.value)
            rows.append(row)
        return rows
    finally:
        wb.release_resources()


class UncachedFormula:
    """Формула XLSX, для которой в файле нет сохранённого результата."""

    def __init__(self, formula: str):
        self.formula = formula


def _validate_xlsx_archive(file_bytes: bytes):
    """Ограничивает распакованный объём XLSX до передачи XML-парсеру."""
    if not zipfile.is_zipfile(io.BytesIO(file_bytes)):
        raise ValueError("файл не является корректным контейнером XLSX")
    with zipfile.ZipFile(io.BytesIO(file_bytes)) as archive:
        entries = archive.infolist()
        if len(entries) > MAX_XLSX_ARCHIVE_ENTRIES:
            raise ValueError(f"в XLSX слишком много внутренних файлов: {len(entries)}")
        unpacked_size = sum(entry.file_size for entry in entries)
        if unpacked_size > MAX_XLSX_UNCOMPRESSED_BYTES:
            raise ValueError(
                "распакованный XLSX превышает допустимый объём "
                f"{MAX_XLSX_UNCOMPRESSED_MB} МБ"
            )


def _read_xlsx_sheet(file_bytes: bytes, sheet_name: str) -> list:
    value_wb = openpyxl.load_workbook(
        io.BytesIO(file_bytes),
        read_only=True,
        data_only=True,
        keep_links=False,
    )
    formula_wb = openpyxl.load_workbook(
        io.BytesIO(file_bytes),
        read_only=True,
        data_only=False,
        keep_links=False,
    )
    try:
        if sheet_name not in value_wb.sheetnames:
            return []
        value_rows = [
            list(row) for row in value_wb[sheet_name].iter_rows(values_only=True)
        ]
        formula_rows = list(formula_wb[sheet_name].iter_rows(values_only=True))
        for row_index, formula_row in enumerate(formula_rows):
            while row_index >= len(value_rows):
                value_rows.append([])
            for column_index, formula_value in enumerate(formula_row):
                if not (
                    isinstance(formula_value, str) and formula_value.startswith("=")
                ):
                    continue
                while column_index >= len(value_rows[row_index]):
                    value_rows[row_index].append(None)
                if value_rows[row_index][column_index] is None:
                    value_rows[row_index][column_index] = UncachedFormula(formula_value)
        return value_rows
    finally:
        formula_wb.close()
        value_wb.close()


def _get_sheet_names(file_bytes: bytes, filename: str) -> list:
    if filename.lower().endswith(".xls"):
        wb = xlrd.open_workbook(file_contents=file_bytes)
        try:
            return wb.sheet_names()
        finally:
            wb.release_resources()
    _validate_xlsx_archive(file_bytes)
    wb = openpyxl.load_workbook(
        io.BytesIO(file_bytes),
        read_only=True,
        data_only=True,
        keep_links=False,
    )
    try:
        return list(wb.sheetnames)
    finally:
        wb.close()


def parse_report(file_bytes: bytes, filename: str) -> dict:
    result = {
        "period": "",
        "daily_by_shift": {},
        "sheets_found": [],
        "used_sheet": "",
        "found_fields": [],
        "report_year": None,
        "report_month": None,
        "warnings": [],
        "error": "",
    }

    try:
        sheet_names = _get_sheet_names(file_bytes, filename)
    except Exception as err:  # noqa: BLE001 - сторонние Excel-парсеры имеют разные исключения
        logger.error(f"Ошибка чтения файла {filename}: {err}")
        result["error"] = f"Не удалось открыть файл как Excel-таблицу: {err}"
        return result

    result["sheets_found"] = sheet_names

    expected_sheet = "ДетСменТекМесяц"
    sheet_lookup = {_normalize_label(name): name for name in sheet_names}
    sheet_name = sheet_lookup.get(_normalize_label(expected_sheet))
    if sheet_name is None:
        result["error"] = (
            f"В файле не найден лист '{expected_sheet}'. "
            f"Доступные листы: {', '.join(sheet_names)}"
        )
        return result
    result["used_sheet"] = sheet_name

    try:
        if filename.lower().endswith(".xls"):
            rows = _read_xls_sheet(file_bytes, sheet_name)
        else:
            rows = _read_xlsx_sheet(file_bytes, sheet_name)
    except Exception as err:  # noqa: BLE001 - повреждённый файл не должен падать из обработчика
        logger.error(f"Ошибка чтения листа {sheet_name}: {err}")
        result["error"] = f"Не удалось прочитать лист '{sheet_name}': {err}"
        return result

    if rows and rows[0] and rows[0][0]:
        result["period"] = str(rows[0][0]).strip()

    date_row_idx = None
    for i, row in enumerate(rows):
        if row and _normalize_label(row[0]) == "дата":
            date_row_idx = i
            break

    if date_row_idx is None:
        result["error"] = f"На листе '{sheet_name}' не найдена строка 'Дата:'."
        return result

    shift_row_idx = date_row_idx + 1
    if (
        shift_row_idx >= len(rows)
        or _normalize_label(rows[shift_row_idx][0]) != "смена"
    ):
        result["error"] = (
            f"На листе '{sheet_name}' не найдена строка 'Смена:' сразу после 'Дата:'."
        )
        return result

    date_row = rows[date_row_idx]
    shift_row = rows[shift_row_idx]

    col_map = {}
    seen_day_shifts = set()
    date_periods = set()
    current_day = None
    for col_i in range(1, max(len(date_row), len(shift_row))):
        d = date_row[col_i] if col_i < len(date_row) else None
        s = shift_row[col_i] if col_i < len(shift_row) else None
        if isinstance(d, (date, datetime)):
            current_day = d.day
            date_periods.add((d.year, d.month))
        elif d is not None and str(d).strip() != "":
            try:
                numeric_day = float(d)
            except (TypeError, ValueError):
                result["error"] = (
                    f"Некорректное значение даты в колонке {col_i + 1}: {d!r}."
                )
                return result
            if not math.isfinite(numeric_day) or not numeric_day.is_integer():
                result["error"] = (
                    f"Некорректное значение даты в колонке {col_i + 1}: {d!r}."
                )
                return result
            current_day = int(numeric_day)

        shift_num = None
        if s is not None and str(s).strip() != "":
            try:
                numeric_shift = float(s)
            except (TypeError, ValueError):
                result["error"] = (
                    f"Некорректный номер смены в колонке {col_i + 1}: {s!r}."
                )
                return result
            if (
                not math.isfinite(numeric_shift)
                or not numeric_shift.is_integer()
                or int(numeric_shift) not in (1, 2)
            ):
                result["error"] = (
                    f"Некорректный номер смены в колонке {col_i + 1}: {s!r}."
                )
                return result
            shift_num = int(numeric_shift)
        if current_day is not None and shift_num in (1, 2):
            if not 1 <= current_day <= 31:
                result["error"] = (
                    f"Некорректный номер дня в колонке {col_i + 1}: {current_day}"
                )
                return result
            day_shift = (current_day, shift_num)
            if day_shift in seen_day_shifts:
                result["error"] = (
                    f"День {current_day}, смена {shift_num} повторяются в нескольких колонках."
                )
                return result
            seen_day_shifts.add(day_shift)
            col_map[col_i] = day_shift

    if len(date_periods) > 1:
        periods = ", ".join(
            f"{month:02d}/{year}" for year, month in sorted(date_periods)
        )
        result["error"] = (
            "В строке дат обнаружено несколько месяцев: "
            f"{periods}. Один файл должен содержать один месячный период."
        )
        return result
    if date_periods:
        result["report_year"], result["report_month"] = next(iter(date_periods))

    if not col_map:
        result["error"] = "Не удалось сопоставить колонки с днями и сменами."
        return result

    daily_by_shift = {}
    found_fields = set()
    for row_idx, row in enumerate(rows[shift_row_idx + 1 :], start=shift_row_idx + 2):
        if not row or not row[0]:
            continue
        name = str(row[0]).strip()
        key = NORMALIZED_CONV_MAP.get(_normalize_label(name))
        if not key:
            continue
        if key in found_fields:
            result["warnings"].append(
                f"Повторная строка «{name}» пропущена (строка {row_idx})."
            )
            continue
        found_fields.add(key)
        for col_i, (day_num, shift_num) in col_map.items():
            if col_i >= len(row):
                continue
            try:
                val = _parse_cell_number(row[col_i])
            except (TypeError, ValueError) as exc:
                result["error"] = (
                    f"Ошибка в строке «{name}», колонка {col_i + 1}: {exc}."
                )
                return result
            daily_by_shift.setdefault(day_num, {1: {}, 2: {}})
            daily_by_shift[day_num][shift_num][key] = val

    missing_bases = sorted(REQUIRED_BASE_FIELDS - found_fields)
    if missing_bases:
        labels = ", ".join(FIELD_LABELS.get(key, key) for key in missing_bases)
        result["error"] = f"Не найдены обязательные строки входных весов: {labels}."
        return result

    warning_groups = (
        (
            "формулах баланса и контроле норм",
            BALANCE_REPORT_FIELDS | NORM_REPORT_FIELDS,
        ),
        ("контроле дублирующих весов", DUPLICATE_REPORT_FIELDS),
        ("расчёте склада концентрата", STOCK_REPORT_FIELDS),
    )
    for purpose, expected_fields in warning_groups:
        missing_fields = sorted(expected_fields - found_fields)
        if not missing_fields:
            continue
        labels = ", ".join(FIELD_LABELS.get(key, key) for key in missing_fields)
        result["warnings"].append(
            f"Не найдены строки, используемые в {purpose}: {labels}. "
            "Для них будет использовано значение 0."
        )

    result["daily_by_shift"] = daily_by_shift
    result["found_fields"] = sorted(found_fields)
    return result


def pct(v: float, base: float) -> float:
    value = finite_number(v)
    denominator = finite_number(base)
    return value / denominator * 100 if denominator else 0.0


def check_norm(val: float, base: float, key: str) -> tuple:
    base = finite_number(base)
    val = finite_number(val)
    if not base or key not in NORMS:
        return "none", 0.0
    p = pct(val, base)
    if val <= 0:
        return "crit", p
    mn, mx = NORMS[key][0], NORMS[key][1]
    margin = (mx - mn) * 0.5
    if mn <= p <= mx:
        return "ok", p
    if mn - margin <= p <= mx + margin:
        return "warn", p
    return "crit", p


def check_doubles(val_main: float, val_dup: float) -> tuple:
    val_main = finite_number(val_main)
    val_dup = finite_number(val_dup)
    if not val_main and not val_dup:
        return "none", 0.0, 0.0
    diff_t = val_main - val_dup
    denominator = max(abs(val_main), abs(val_dup))
    diff_p = abs(diff_t) / denominator * 100
    if diff_p <= DUP_WARN:
        return "ok", diff_p, diff_t
    if diff_p <= DUP_CRIT:
        return "warn", diff_p, diff_t
    return "crit", diff_p, diff_t


# ════════════════════════════════════════════════════════
#  ФОРМАТИРОВАНИЕ
# ════════════════════════════════════════════════════════
def em_bal(v: float | None) -> str:
    if v is None or not math.isfinite(v):
        return "⬜"
    a = abs(v)
    return "✅" if a <= WARN_PCT else "⚠️" if a <= CRIT_PCT else "🚨"


def em_norm(st: str) -> str:
    return {"ok": "✅", "warn": "⚠️", "crit": "🚨", "none": "⬜"}.get(st, "⬜")


def em_dup(st: str) -> str:
    return {"ok": "✅", "warn": "⚠️", "crit": "🚨", "none": "⬜"}.get(st, "⬜")


def sign(v: float | None, d: int = 2) -> str:
    if v is None or not math.isfinite(v):
        return "—"
    return f"+{v:.{d}f}%" if v >= 0 else f"{v:.{d}f}%"


def fmt(v) -> str:
    if v is None:
        return "—"
    number = finite_number(v, default=math.nan)
    if not math.isfinite(number):
        return "—"
    return f"{round(number):,}".replace(",", " ")


def fmt2(v) -> str:
    if v is None:
        return "—"
    number = finite_number(v, default=math.nan)
    if not math.isfinite(number):
        return "—"
    return f"{number / 1000:.1f}k" if abs(number) >= 1000 else f"{number:.0f}"


def build_alerts(d: dict, label: str = "", include_balances: bool = True) -> list:
    alerts = []
    base4, base3 = d.get("kv4", 0), d.get("kv3", 0)
    prefix = f"[{label}] " if label else ""

    if include_balances:
        for name, val in [
            ("Баланс 1", bal1(d)),
            ("Баланс 2", bal2(d)),
            ("Баланс С.1", balc1(d)),
            ("Баланс С.2", balc2(d)),
        ]:
            if val is None:
                continue
            if abs(val) > CRIT_PCT:
                alerts.append(
                    (
                        "crit",
                        f"🚨 {prefix}{name} = {sign(val)} (критично, норма ±{CRIT_PCT:.0f}%)",
                    )
                )
            elif abs(val) > WARN_PCT:
                alerts.append(
                    ("warn", f"⚠️ {prefix}{name} = {sign(val)} (норма ±{WARN_PCT:.0f}%)")
                )

    checks = [
        ("kv14", base4),
        ("kv32", base4),
        ("kv34", base4),
        ("kv34a", base4),
        ("kv102", base4),
        ("kv15", base3),
        ("kv19", base3),
        ("kv31", base3),
        ("kv33", base3),
        ("kv101", base3),
    ]
    for key, base in checks:
        val = d.get(key, 0)
        if not finite_number(base):
            continue
        st, p = check_norm(val, base, key)
        if st in ("warn", "crit"):
            mn, mx, desc, _ = NORMS[key]
            alerts.append(
                (st, f"{em_norm(st)} {prefix}{desc}: {p:.1f}% (норма {mn}–{mx}%)")
            )

    st4, p4, t4 = check_doubles(base4, d.get("kv4d", 0))
    if st4 in ("warn", "crit"):
        alerts.append(
            (
                st4,
                f"{em_dup(st4)} {prefix}Конв.4 vs 4Д: расхождение {p4:.2f}% ({fmt(t4)} т)",
            )
        )
    st3, p3, t3 = check_doubles(base3, d.get("kv3d", 0))
    if st3 in ("warn", "crit"):
        alerts.append(
            (
                st3,
                f"{em_dup(st3)} {prefix}Конв.3 vs 3Д: расхождение {p3:.2f}% ({fmt(t3)} т)",
            )
        )

    return alerts


def build_rolling_balance_alerts(rows: list) -> list:
    """Сигналы баланса только по трём последовательным завершённым суткам."""
    snapshots = rolling_snapshots(rows, periods=(3,), fields=FIELDS)
    if not snapshots:
        return []

    snapshot = snapshots[0]
    if not snapshot["consecutive"]:
        return []

    result = []
    labels = (
        ("Баланс 1", "b1"),
        ("Баланс С.1", "bc1"),
        ("Баланс 2", "b2"),
        ("Баланс С.2", "bc2"),
    )
    for name, key in labels:
        value = snapshot["balances"][key]
        status = balance_status(value, WARN_PCT, CRIT_PCT)
        if status == "crit":
            result.append(
                (
                    "crit",
                    (
                        f"🚨 [3 суток {snapshot['label']}] {name} = {sign(value)} "
                        f"(устойчивое критичное отклонение, более ±{CRIT_PCT:.0f}%)"
                    ),
                )
            )
        elif status == "warn":
            result.append(
                (
                    "warn",
                    (
                        f"⚠️ [3 суток {snapshot['label']}] {name} = {sign(value)} "
                        f"(устойчивое отклонение, более ±{WARN_PCT:.0f}%)"
                    ),
                )
            )
    return result


# ════════════════════════════════════════════════════════
#  AI АГЕНТ
# ════════════════════════════════════════════════════════
SYSTEM_PROMPT = """Ты — AI-агент метролога горно-обогатительной фабрики (ДОФ), Костанайский регион, Казахстан.

ТЕХНОЛОГИЧЕСКАЯ СХЕМА:
Очередь 1: Руда → Конв.4 (6400 т/ч) → 6 бункеров (6000т каждый) → дробление/грохочение
→ Конв.14 (2500 т/ч) → 5 бункеров (1000т) → Сепарация →
  [ПП]: Конв.18(условный)→Конв.32(2500 т/ч)→Конв.34(3200 т/ч)→ММО
  [Хвосты]: Конв.20(условный)→Конв.24(2500 т/ч)→Склад хв.№25  ИЛИ  Конв.102(800 т/ч)→Склад хв.№104
  [Подрешетный продукт 1 стадии грохочения]: Конв.10/34А(1250 т/ч)→Конв.32→Конв.34
  [Склад ПП]: Конв.24→Склад ПП; Склад ПП→Конв.28А.I→Конв.32→Конв.34
НЕ в балансе: Конв.18, Конв.20

Очередь 2: Руда → Конв.3 (6400 т/ч) → 6 бункеров (6000т) → дробление/грохочение
→ Конв.15 (2500 т/ч) → 5 бункеров (1000т) → Сепарация →
  [ПП]: Конв.19(2500 т/ч)→Конв.31→Конв.33→ММО
  [Хвосты]: Конв.101(800 т/ч)→Склад хв.№104
  [Склад ПП]: Конв.28А.II→Конв.31→Конв.33

НОРМЫ БАЛАНСА (% от Конв.3 или Конв.4 = 100%):
• Конв.14/15: 70% ±10% (60–80%) — задержка руды в бункерах объясняет расхождение, это НОРМАЛЬНО
• Конв.101/102: 13% ±7% (6–20%) — хвосты сепарации
• Конв.19/31/32: 50% ±10% (40–60%) — промпродукт
• Конв.33/34: 83–87% — промпродукт + подрешетный продукт + склад ПП → ММО
• Конв.10/34А: 30% ±15% (15–45%) — подрешетный продукт 1 стадии грохочения

ФОРМУЛЫ БАЛАНСОВ:
Баланс1  = 102+34+24П+24Хв−28А.I−4   (% от Конв.4)
БалансС1 = 102+24Хв+24П+32−28А.I−14
Баланс2  = 101+33−28А.II−3            (% от Конв.3)
БалансС2 = 101+31−28А.II−15

ПОРОГИ: ±2% предупреждение, ±5% критично
Дублирующие весы (Конв.4/4Д, Конв.3/3Д): норма расхождения <0.5%, критично >2%
Для устойчивого сигнала материального баланса суммируй тоннаж за 3 последние
последовательные завершённые сутки и только затем рассчитывай процент.

Никогда не выдумывай отсутствующие показания и не выдавай возможную причину за
установленный факт. Чётко разделяй: (1) что прямо следует из цифр; (2) гипотезы;
(3) каких данных не хватает для подтверждения. Межсуточный переход материала
можно предположить по смене знака суточного отклонения и нормализации окна 2–3
суток, но нельзя доказать без почасовых/посменных показаний и остатков в бункерах.
Давай конкретные метрологические рекомендации (поверка, чистка датчиков,
проверка калибровки, осмотр ленты), помечая их как рекомендации.
Отвечай по-русски, кратко, для Telegram. Используй ✅⚠️🚨🔧📊🔮."""


async def ask_ai(question: str, context: str) -> str:
    if not ANTHROPIC_API_KEY:
        return "⚠️ AI недоступен: не задан ANTHROPIC_API_KEY на сервере."

    payload = {
        "model": AI_MODEL,
        "max_tokens": 1200,
        "system": SYSTEM_PROMPT,
        "messages": [{"role": "user", "content": f"{question}\n\nДанные:\n{context}"}],
    }
    timeout = aiohttp.ClientTimeout(total=45, connect=15)
    try:
        async with (
            aiohttp.ClientSession(timeout=timeout) as session,
            session.post(
                "https://api.anthropic.com/v1/messages",
                headers={
                    "x-api-key": ANTHROPIC_API_KEY,
                    "anthropic-version": "2023-06-01",
                    "content-type": "application/json",
                },
                json=payload,
            ) as response,
        ):
            if response.status != 200:
                body = await response.text()
                logger.error("Anthropic API error %s: %s", response.status, body[:1000])
                return f"⚠️ Ошибка AI-сервера (код {response.status})"
            data = await response.json()
            if not isinstance(data, dict) or not isinstance(data.get("content"), list):
                logger.error("Anthropic API returned unexpected JSON: %r", data)
                return "⚠️ AI-сервер вернул ответ неизвестного формата."
            text_parts = [
                block.get("text", "")
                for block in data.get("content", [])
                if isinstance(block, dict)
                and block.get("type") == "text"
                and block.get("text")
            ]
            if not text_parts:
                logger.error("Anthropic API returned no text blocks: %r", data)
                return "⚠️ AI-сервер вернул ответ без текста."
            answer = "\n".join(text_parts)
            if data.get("stop_reason") == "max_tokens":
                answer += "\n\n⚠️ Ответ остановлен по лимиту длины; уточните вопрос."
            return answer
    except asyncio.TimeoutError:
        logger.error("Anthropic API timeout")
        return "⚠️ AI-сервер не ответил за 45 секунд. Повторите запрос позже."
    except (aiohttp.ClientError, ValueError, KeyError) as exc:
        logger.error("AI connection error: %s", exc)
        return "⚠️ Не удалось получить корректный ответ AI-сервера."


AI_CONTEXT_FIELDS = (
    ("К4", "kv4"),
    ("К4Д", "kv4d"),
    ("К14", "kv14"),
    ("К32", "kv32"),
    ("К34", "kv34"),
    ("К10/34А", "kv34a"),
    ("К102", "kv102"),
    ("К24ПП", "kv24p"),
    ("К24ХВ", "kv24hv"),
    ("К28А.I", "kv28a1"),
    ("К3", "kv3"),
    ("К3Д", "kv3d"),
    ("К15", "kv15"),
    ("К19", "kv19"),
    ("К31", "kv31"),
    ("К33", "kv33"),
    ("К101", "kv101"),
    ("К28А.II", "kv28a2"),
)


def _ai_tonnage_line(data: dict) -> str:
    return "; ".join(
        f"{label}={finite_number(data.get(key)):.2f} т"
        for label, key in AI_CONTEXT_FIELDS
    )


def make_ai_context(rows: list, rolling_rows: list | None = None) -> str:
    full_days = sorted(
        rows,
        key=lambda row: (row["year"], row["month"], row["day_num"]),
    )
    if not full_days:
        return "Нет завершённых суток с данными."

    s = sum_period(full_days, FIELDS)
    lines = [
        "ФАКТИЧЕСКИЕ ДАННЫЕ ИЗ БАЗЫ БОТА. Все тоннажи абсолютные.",
        (
            f"Активный период: {format_period_label(full_days)}; "
            f"завершённых суток: {len(full_days)}."
        ),
        "ИТОГ АКТИВНОГО ПЕРИОДА:",
        _ai_tonnage_line(s),
        (
            f"Б1={sign(bal1(s))}; БС1={sign(balc1(s))}; "
            f"Б2={sign(bal2(s))}; БС2={sign(balc2(s))}."
        ),
        "",
        "ПОСУТОЧНЫЕ АБСОЛЮТНЫЕ ПОКАЗАНИЯ:",
    ]

    for row in full_days:
        report_day = date(row["year"], row["month"], row["day_num"])
        lines.extend(
            [
                f"Дата {report_day:%d.%m.%Y}: {_ai_tonnage_line(row)}",
                (
                    f"Б1={sign(bal1(row))}; БС1={sign(balc1(row))}; "
                    f"Б2={sign(bal2(row))}; БС2={sign(balc2(row))}."
                ),
            ]
        )

    window_rows = rolling_rows if rolling_rows is not None else full_days[-3:]
    snapshots = rolling_snapshots(
        window_rows,
        periods=(1, 2, 3),
        fields=FIELDS,
    )
    lines.extend(
        ["", "СКОЛЬЗЯЩИЕ ОКНА (сначала сложен тоннаж, затем рассчитан процент):"]
    )
    for snapshot in snapshots:
        balances = snapshot["balances"]
        lines.extend(
            [
                (
                    f"Окно {snapshot['days']} сут.; {snapshot['label']}; "
                    f"последовательные даты: "
                    f"{'да' if snapshot['consecutive'] else 'нет'}."
                ),
                _ai_tonnage_line(snapshot["total"]),
                (
                    f"Б1={sign(balances['b1'])}; БС1={sign(balances['bc1'])}; "
                    f"Б2={sign(balances['b2'])}; БС2={sign(balances['bc2'])}."
                ),
            ]
        )

    lines.extend(["", "УСТОЙЧИВЫЙ СИГНАЛ БАЛАНСА ПО 3 СУТКАМ:"])
    stable_alerts = build_rolling_balance_alerts(window_rows)
    if stable_alerts:
        lines.extend(alert[1] for alert in stable_alerts)
    elif len(window_rows) < 3:
        lines.append("Недостаточно завершённых суток для сигнала.")
    elif not is_consecutive_period(window_rows[-3:]):
        lines.append("Сигнал не сформирован из-за пропуска дат.")
    else:
        lines.append("Устойчивого отклонения не обнаружено.")

    lines.extend(["", "СУТОЧНЫЕ НАРУШЕНИЯ НОРМ КОНВЕЙЕРОВ И ДУБЛЕЙ:"])
    found = False
    for r in full_days:
        al = build_alerts(r, label=f"д.{r['day_num']}", include_balances=False)
        if al:
            found = True
            lines.extend(a[1] for a in al)
    if not found:
        lines.append("Нарушений не обнаружено.")
    return "\n".join(lines)


# ════════════════════════════════════════════════════════
#  TELEGRAM BOT
# ════════════════════════════════════════════════════════
bot = Bot(token=BOT_TOKEN)
dp = Dispatcher(storage=MemoryStorage())
REPORT_IMPORT_LOCK = asyncio.Lock()


class AccessMiddleware(BaseMiddleware):
    async def __call__(self, handler, event: Message, data):
        user = event.from_user
        if user is None:
            await event.answer(
                "⛔ Команды бота недоступны для сообщений от имени канала."
            )
            return None
        if ALLOWED_USER_IDS and user.id not in ALLOWED_USER_IDS:
            await event.answer("⛔ У вас нет доступа к этому боту.")
            return None
        return await handler(event, data)


dp.message.middleware(AccessMiddleware())


class AIState(StatesGroup):
    waiting_for_question = State()


class StockInput(StatesGroup):
    waiting_for_prev = State()
    waiting_for_curr = State()


class StockNightInput(StatesGroup):
    waiting_for_night = State()
    waiting_for_morning = State()


def md_escape(value) -> str:
    """Экранирование динамического текста для Telegram legacy Markdown."""
    text = str(value or "")
    for char in ("\\", "`", "*", "_", "["):
        text = text.replace(char, f"\\{char}")
    return text


def split_message(text: str, limit: int = 3900) -> list[str]:
    """Делит сообщение по строкам, не разрезая Markdown посередине строки."""
    if len(text) <= limit:
        return [text]

    chunks: list[str] = []
    current = ""
    for line in text.splitlines(keepends=True):
        if len(line) > limit:
            if current:
                chunks.append(current.rstrip())
                current = ""
            chunks.extend(
                line[index : index + limit] for index in range(0, len(line), limit)
            )
            continue
        if current and len(current) + len(line) > limit:
            chunks.append(current.rstrip())
            current = line
        else:
            current += line
    if current:
        chunks.append(current.rstrip())
    return [chunk for chunk in chunks if chunk]


async def answer_markdown(msg: Message, text: str):
    for chunk in split_message(text):
        await msg.answer(chunk, parse_mode="Markdown")


def parse_user_number(text: str | None) -> float:
    if text is None:
        raise ValueError("нет текста")
    normalized = text.replace("\xa0", " ").strip().replace(" ", "").replace(",", ".")
    value = float(normalized)
    if not math.isfinite(value):
        raise ValueError("число должно быть конечным")
    return value


def day_label(year: int, month: int, day_num: int) -> str:
    return date(year, month, day_num).strftime("%d.%m.%Y")


def previous_day_label(year: int, month: int, day_num: int) -> str:
    return (date(year, month, day_num) - timedelta(days=1)).strftime("%d.%m.%Y")


def _is_cancel_command(msg: Message) -> bool:
    """[FIX]: Утилита защиты от залипания FSM. Определяет нажатия кнопок меню."""
    text = msg.text
    if not text:
        return True  # Если прислали фото/файл — отменяем ввод
    # Проверяем, начинается ли текст с эмодзи кнопок или слэша команды
    prefixes = (
        "📅",
        "🌙",
        "📆",
        "🗓",
        "📉",
        "🔔",
        "🔍",
        "🏭",
        "📥",
        "🌅",
        "❓",
        "🤖",
        "/",
    )
    return any(text.startswith(p) for p in prefixes)


def main_keyboard() -> ReplyKeyboardMarkup:
    return ReplyKeyboardMarkup(
        keyboard=[
            [
                KeyboardButton(text="📅 Суточный баланс"),
                KeyboardButton(text="🌙 Ночная смена"),
            ],
            [
                KeyboardButton(text="📆 Недельная сводка"),
                KeyboardButton(text="🗓 Месячный итог"),
            ],
            [KeyboardButton(text="📉 Просмотр проскальзывания")],
            [
                KeyboardButton(text="🔔 Алерты"),
                KeyboardButton(text="🔍 Дублирование весов"),
            ],
            [
                KeyboardButton(text="🏭 Склад концентрата"),
                KeyboardButton(text="📥 Ввести запас"),
            ],
            [KeyboardButton(text="🌅 Склад — ночная смена")],
            [KeyboardButton(text="❓ Помощь"), KeyboardButton(text="🤖 Спросить AI")],
        ],
        resize_keyboard=True,
    )


@dp.message(Command("start"))
async def cmd_start(msg: Message, state: FSMContext):
    await state.clear()
    await asyncio.to_thread(init_db)
    await msg.answer(
        "⚖️ *ДОФ Баланс*\n"
        "Мониторинг конвейерных весов · 2 очереди производства\n\n"
        "📁 Просто отправьте файл отчёта *report.xls* — бот автоматически "
        "прочитает данные, посчитает балансы и проверит технологические нормы.",
        parse_mode="Markdown",
        reply_markup=main_keyboard(),
    )


@dp.message(F.document)
async def handle_report(msg: Message, state: FSMContext):
    await state.clear()
    doc: Document = msg.document
    fn = doc.file_name or ""
    if not (fn.lower().endswith(".xls") or fn.lower().endswith(".xlsx")):
        await msg.answer("❌ Поддерживаются только файлы .xls или .xlsx")
        return
    if doc.file_size and doc.file_size > MAX_REPORT_BYTES:
        await msg.answer(
            f"❌ Файл слишком большой: допустимо не более {MAX_REPORT_SIZE_MB} МБ."
        )
        return

    wait_msg = await msg.answer("⏳ Читаю файл...")

    try:
        file_obj = await bot.get_file(doc.file_id)
        buf = io.BytesIO()
        await bot.download_file(file_obj.file_path, buf)
        file_bytes = buf.getvalue()
    except Exception as exc:  # noqa: BLE001 - исключения загрузчика зависят от aiogram
        # В сетевом traceback потенциально может оказаться URL Telegram с токеном.
        logger.error("Не удалось скачать Telegram-файл %r (%s)", fn, type(exc).__name__)
        await wait_msg.edit_text(
            "❌ Не удалось скачать файл из Telegram. Повторите отправку позже."
        )
        return
    if len(file_bytes) > MAX_REPORT_BYTES:
        await wait_msg.edit_text(
            f"❌ Файл слишком большой: допустимо не более {MAX_REPORT_SIZE_MB} МБ."
        )
        return

    parsed = await asyncio.to_thread(parse_report, file_bytes, fn)

    if parsed.get("error") or not parsed["daily_by_shift"]:
        err = parsed.get("error", "Неизвестная ошибка структуры файла.")
        await wait_msg.edit_text(f"⚠️ {err}")
        return

    now = local_now()
    report_year, report_month, period_detected = infer_report_period(
        f"{parsed.get('period', '')} {fn}",
        now,
    )
    dates_year = parsed.get("report_year")
    dates_month = parsed.get("report_month")
    if dates_year is not None and dates_month is not None:
        dates_period = (int(dates_year), int(dates_month))
        if period_detected and (report_year, report_month) != dates_period:
            await wait_msg.edit_text(
                "⚠️ Отчёт не сохранён: период в заголовке или имени файла "
                f"({report_month:02d}/{report_year}) не совпадает с датами "
                f"в таблице ({dates_period[1]:02d}/{dates_period[0]})."
            )
            return
        report_year, report_month = dates_period
        period_detected = True
    try:
        async with REPORT_IMPORT_LOCK:
            saved, max_day, incomplete_day = await asyncio.to_thread(
                db_save_daily,
                parsed,
                msg.from_user.id,
                fn,
                report_year,
                report_month,
                now.date(),
            )
    except ReportDataError as exc:
        await wait_msg.edit_text(f"⚠️ Отчёт не сохранён: {exc}")
        return
    except sqlite3.Error:
        logger.exception("Ошибка базы при импорте %r", fn)
        await wait_msg.edit_text(
            "❌ Не удалось сохранить отчёт в базу. Старые данные не изменены."
        )
        return

    night_available = False
    if incomplete_day is not None:
        night_available = (
            await asyncio.to_thread(db_get_night_shift, report_year, report_month)
            is not None
        )

    if saved == 0:
        if incomplete_day is not None and night_available:
            text = (
                "🌙 Файл прочитан. Текущий день ещё не завершён, поэтому он не включён "
                "в суточный и скользящий баланс. Данные первой смены доступны по кнопке "
                "*🌙 Ночная смена*."
            )
        elif incomplete_day is not None:
            text = (
                "🌙 Файл прочитан. Текущий день ещё не завершён и не включён "
                "в суточный баланс; данных первой смены в файле нет."
            )
        else:
            text = "⚠️ Файл прочитан, но завершённые сутки с данными не найдены."
        await wait_msg.edit_text(text, parse_mode="Markdown")
        return

    period_source = (
        "определён из отчёта или имени файла"
        if period_detected
        else "взят по дате загрузки"
    )
    if incomplete_day is not None:
        if night_available:
            last_day_note = (
                f"🌙 День {incomplete_day}: исключён из суточного и скользящего "
                f"баланса целиком (Смена 2 ещё идёт) — Смена 1 доступна "
                f"отдельно по кнопке"
            )
        else:
            last_day_note = (
                f"🌙 День {incomplete_day}: исключён из суточного и скользящего "
                f"баланса; данных Смены 1 в файле нет"
            )
    else:
        last_day_note = (
            f"✅ Все найденные дни, включая день {max_day}, сохранены как завершённые"
        )

    warning_text = ""
    if parsed.get("warnings"):
        warning_lines = "\n".join(f"• {md_escape(item)}" for item in parsed["warnings"])
        warning_text = f"\n\n⚠️ *Замечания к структуре:*\n{warning_lines}"

    await wait_msg.edit_text(
        f"✅ *Отчёт загружен!*\n\n"
        f"📊 Лист: `{parsed['used_sheet']}`\n"
        f"📄 Файл: {md_escape(fn)}\n"
        f"📅 Период: {md_escape(parsed.get('period') or 'текущий месяц')}\n"
        f"🗃 Период базы: {report_month:02d}/{report_year} ({period_source})\n"
        f"💾 В баланс включено полных суток: {saved}\n"
        f"{last_day_note}"
        f"{warning_text}\n\n"
        f"Нажмите *📉 Просмотр проскальзывания* для расчёта за 1, 2 и 3 суток.",
        parse_mode="Markdown",
    )


# ── СУТОЧНЫЙ БАЛАНС ───────────────────────────────────────
@dp.message(F.text == "📅 Суточный баланс")
async def report_daily(msg: Message, state: FSMContext):
    await state.clear()
    active = db_get_active_report(msg.from_user.id)
    if not active:
        await msg.answer("📭 Сначала загрузите свежий файл отчёта.")
        return
    year, month = active["year"], active["month"]
    rows = db_get_active_month_data(active)
    if not rows:
        await msg.answer("📭 Нет завершённых суток. Загрузите файл отчёта.")
        return

    lines = [
        f"📊 *Суточный баланс ({month:02d}/{year})*",
        (
            f"_Последняя загрузка: "
            f"{md_escape(active.get('filename') or 'отчёт')}; "
            f"только завершённые сутки_\n"
        ),
    ]
    lines.append("`Дн  Б1      Б2      Кв4     Кв3`")
    for r in rows:
        b1, b2 = bal1(r), bal2(r)
        lines.append(
            f"`{r['day_num']:>2d}  "
            f"{em_bal(b1)}{sign(b1, 1):>6s}  "
            f"{em_bal(b2)}{sign(b2, 1):>6s}  "
            f"{fmt2(r.get('kv4')):>6s}  "
            f"{fmt2(r.get('kv3')):>6s}`"
        )
    await answer_markdown(msg, "\n".join(lines))


# ── НОЧНАЯ СМЕНА ──────────────────────────────────────────
@dp.message(F.text == "🌙 Ночная смена")
async def report_night_shift(msg: Message, state: FSMContext):
    await state.clear()
    active = db_get_active_report(msg.from_user.id)
    if not active:
        await msg.answer("📋 Сначала загрузите свежий report.xls.")
        return
    year, month = active["year"], active["month"]
    ns = db_get_night_shift(year, month)
    if not ns:
        await msg.answer("📋 Нет данных по ночной смене.\nЗагрузите свежий report.xls.")
        return

    day_num = ns["day_num"]
    b1, b2 = bal1(ns), bal2(ns)
    bc1, bc2 = balc1(ns), balc2(ns)

    text = (
        f"🌙 *Ночная смена — день {day_num:02d}.{month:02d}*\n"
        f"_Смена 1 (19:30–07:30), отдельно от месячного баланса_\n\n"
        f"━━ 1 очередь (Кв4={fmt(ns.get('kv4'))}т) ━━\n"
        f"{em_bal(b1)} Баланс 1: {sign(b1)}\n"
        f"{em_bal(bc1)} Баланс С.1: {sign(bc1)}\n"
        f"Конв.14: {fmt(ns.get('kv14'))}т  Конв.102: {fmt(ns.get('kv102'))}т\n\n"
        f"━━ 2 очередь (Кв3={fmt(ns.get('kv3'))}т) ━━\n"
        f"{em_bal(b2)} Баланс 2: {sign(b2)}\n"
        f"{em_bal(bc2)} Баланс С.2: {sign(bc2)}\n"
        f"Конв.15: {fmt(ns.get('kv15'))}т  Конв.101: {fmt(ns.get('kv101'))}т\n\n"
        f"_Смена 2 этого дня ещё идёт — будет учтена завтра, "
        f"когда сутки станут полными._"
    )
    await msg.answer(text, parse_mode="Markdown")


# ── СКОЛЬЗЯЩИЙ БАЛАНС ────────────────────────────────────
@dp.message(F.text == "📉 Скользящий баланс")
@dp.message(F.text == "📉 Просмотр проскальзывания")
@dp.message(F.text == "Просмотр проскальзывания")
async def report_sliding_balance(msg: Message, state: FSMContext):
    await state.clear()
    active = db_get_active_report(msg.from_user.id)
    end_date = active_report_end_date(active)
    if not active:
        await msg.answer("📭 Сначала загрузите свежий файл отчёта.")
        return
    rows = db_get_latest_completed_days(3, end_date=end_date) if end_date else []
    if not rows:
        await msg.answer("📭 Нет завершённых суток. Загрузите файл отчёта.")
        return

    snapshots = rolling_snapshots(rows, periods=(1, 2, 3), fields=FIELDS)
    captions = {1: "1 сутки", 2: "2 суток", 3: "3 суток"}
    lines = [
        "📉 *Скользящий баланс*",
        "_Суммируется тоннаж завершённых суток; текущая ночная смена не включается._\n",
    ]

    for snapshot in snapshots:
        balances = snapshot["balances"]
        lines.extend(
            [
                f"*{captions[snapshot['days']]} · {snapshot['label']}*",
                (
                    f"1 очередь: {em_bal(balances['b1'])} Б1 {sign(balances['b1'])}   "
                    f"{em_bal(balances['bc1'])} БС.1 {sign(balances['bc1'])}"
                ),
                (
                    f"2 очередь: {em_bal(balances['b2'])} Б2 {sign(balances['b2'])}   "
                    f"{em_bal(balances['bc2'])} БС.2 {sign(balances['bc2'])}"
                ),
                "",
            ]
        )
        if snapshot["days"] > 1 and not snapshot["consecutive"]:
            lines.extend(
                [
                    (
                        "⚠️ В этом окне есть пропуск календарных дат; "
                        "значение приведено только справочно."
                    ),
                    "",
                ]
            )

    snapshot_3d = next((item for item in snapshots if item["days"] == 3), None)
    if snapshot_3d is None:
        lines.append(
            f"ℹ️ Для устойчивого сигнала нужны 3 завершённых суток; сейчас доступно {len(rows)}."
        )
    elif not snapshot_3d["consecutive"]:
        lines.append(
            "⚠️ Между последними записями есть пропуск дат. Значения показаны, "
            "но устойчивый трёхсуточный сигнал не формируется."
        )
    else:
        stable_alerts = build_rolling_balance_alerts(rows)
        if stable_alerts:
            lines.append("*Устойчивый сигнал по 3 суткам:*")
            lines.extend(alert[1] for alert in stable_alerts)
        else:
            lines.append("✅ Устойчивого отклонения баланса за 3 суток не обнаружено.")

    await answer_markdown(msg, "\n".join(lines))


# ── НЕДЕЛЬНАЯ СВОДКА ─────────────────────────────────────
@dp.message(F.text == "📆 Недельная сводка")
async def report_weekly(msg: Message, state: FSMContext):
    await state.clear()
    active = db_get_active_report(msg.from_user.id)
    end_date = active_report_end_date(active)
    if not active:
        await msg.answer("📭 Сначала загрузите свежий файл отчёта.")
        return
    rows = db_get_latest_completed_days(7, end_date=end_date) if end_date else []
    if not rows:
        await msg.answer("📭 Нет завершённых суток.")
        return

    last7 = rows[-7:]
    s = sum_period(last7, FIELDS)
    lines = [
        f"📆 *Сводка за {len(last7)} последних суток*",
        f"_{format_period_label(last7)}_\n",
    ]
    lines.append(f"{em_bal(bal1(s))} Баланс 1: {sign(bal1(s))}")
    lines.append(f"{em_bal(bal2(s))} Баланс 2: {sign(bal2(s))}")
    lines.append(f"Конв.4: {fmt(s['kv4'])}т   Конв.3: {fmt(s['kv3'])}т\n")
    if len(last7) > 1 and not is_consecutive_period(last7):
        lines.append(
            "⚠️ Между записями есть пропуск календарных дат; итог рассчитан "
            "по имеющимся завершённым суткам.\n"
        )
    for r in last7:
        lines.append(
            f"▪️ {r['day_num']:02d}.{r['month']:02d}: "
            f"Кв4={fmt(r.get('kv4'))}т  Кв3={fmt(r.get('kv3'))}т"
        )
    await answer_markdown(msg, "\n".join(lines))


# ── МЕСЯЧНЫЙ ИТОГ ─────────────────────────────────────────
@dp.message(F.text == "🗓 Месячный итог")
async def report_monthly(msg: Message, state: FSMContext):
    await state.clear()
    active = db_get_active_report(msg.from_user.id)
    if not active:
        await msg.answer("📭 Сначала загрузите свежий файл отчёта.")
        return
    year, month = active["year"], active["month"]
    rows = db_get_active_month_data(active)
    if not rows:
        await msg.answer("📭 Нет полных данных за месяц.")
        return

    s = sum_period(rows, FIELDS)
    b1, b2 = bal1(s), bal2(s)
    bc1, bc2 = balc1(s), balc2(s)

    text = (
        f"🗓 *Месячный итог ({month:02d}/{year})*\n"
        f"_Учтено суток: {len(rows)}, текущая смена исключена_\n\n"
        f"━━ 1 очередь ━━\n"
        f"Конв.4: {fmt(s['kv4'])}т (осн.)   Конв.4Д: {fmt(s['kv4d'])}т (дубл.)\n"
        f"Конв.14: {fmt(s['kv14'])}т  ({pct(s['kv14'], s['kv4']):.0f}%)\n"
        f"Конв.32: {fmt(s['kv32'])}т  ({pct(s['kv32'], s['kv4']):.0f}%)\n"
        f"Конв.34: {fmt(s['kv34'])}т  ({pct(s['kv34'], s['kv4']):.0f}%)\n"
        f"Конв.102: {fmt(s['kv102'])}т  ({pct(s['kv102'], s['kv4']):.0f}%)\n"
        f"{em_bal(b1)} Баланс 1: {sign(b1)}\n"
        f"{em_bal(bc1)} Баланс С.1: {sign(bc1)}\n\n"
        f"━━ 2 очередь ━━\n"
        f"Конв.3: {fmt(s['kv3'])}т (осн.)   Конв.3Д: {fmt(s['kv3d'])}т (дубл.)\n"
        f"Конв.15: {fmt(s['kv15'])}т  ({pct(s['kv15'], s['kv3']):.0f}%)\n"
        f"Конв.33: {fmt(s['kv33'])}т  ({pct(s['kv33'], s['kv3']):.0f}%)\n"
        f"Конв.101: {fmt(s['kv101'])}т  ({pct(s['kv101'], s['kv3']):.0f}%)\n"
        f"{em_bal(b2)} Баланс 2: {sign(b2)}\n"
        f"{em_bal(bc2)} Баланс С.2: {sign(bc2)}"
    )
    await msg.answer(text, parse_mode="Markdown")


# ── АЛЕРТЫ ────────────────────────────────────────────────
@dp.message(F.text == "🔔 Алерты")
async def report_alerts(msg: Message, state: FSMContext):
    await state.clear()
    active = db_get_active_report(msg.from_user.id)
    if not active:
        await msg.answer("📭 Сначала загрузите свежий файл отчёта.")
        return
    end_date = active_report_end_date(active)
    rows = db_get_active_month_data(active)
    rolling_rows = (
        db_get_latest_completed_days(3, end_date=end_date) if end_date else []
    )
    if not rows and not rolling_rows:
        await msg.answer("📭 Нет завершённых суток.")
        return

    # Суточные нормы конвейеров и дубли проверяются по каждому дню.
    # Сам материальный баланс сигнализирует только по устойчивому окну 3 суток.
    all_alerts = build_rolling_balance_alerts(rolling_rows)
    for r in rows:
        all_alerts.extend(
            build_alerts(
                r,
                label=f"{r['day_num']:02d}.{r['month']:02d}",
                include_balances=False,
            )
        )

    rolling_notice = ""
    if len(rolling_rows) < 3:
        rolling_notice = (
            f"ℹ️ Трёхсуточный сигнал баланса пока не рассчитан: "
            f"доступно {len(rolling_rows)} завершённых суток из 3."
        )
    elif not is_consecutive_period(rolling_rows):
        rolling_notice = (
            "⚠️ Трёхсуточный сигнал баланса не сформирован: "
            "между последними записями есть пропуск дат."
        )

    stock_history = db_get_active_stock_history(active)
    stock_alerts = []
    for r in stock_history:
        stock_curr = r.get("stock_curr", 0)
        nesovp = r.get("nesovpadenie", 0)
        day = r["day_num"]
        if stock_curr < 0:
            stock_alerts.append(
                ("crit", f"🚨 Склад д.{day}: запас отрицательный ({fmt(stock_curr)}т)!")
            )
        elif abs(nesovp) > STOCK_DIFF_WARN:
            em = "crit" if abs(nesovp) > STOCK_DIFF_WARN * 2 else "warn"
            stock_alerts.append(
                (
                    em,
                    f"{'🚨' if em == 'crit' else '⚠️'} Склад д.{day}: несовпадение {fmt(nesovp)}т",
                )
            )

    has_any = all_alerts or stock_alerts
    if not has_any:
        text = "✅ *Нарушений не обнаружено за весь период.*"
        if rolling_notice:
            text += f"\n\n{rolling_notice}"
        await msg.answer(text, parse_mode="Markdown")
        return

    crits = [a[1] for a in all_alerts if a[0] == "crit"]
    warns = [a[1] for a in all_alerts if a[0] == "warn"]
    s_crits = [a[1] for a in stock_alerts if a[0] == "crit"]
    s_warns = [a[1] for a in stock_alerts if a[0] == "warn"]

    total_c = len(crits) + len(s_crits)
    total_w = len(warns) + len(s_warns)
    lines = [f"🔔 *Алерты: {total_c} крит. / {total_w} предупр.*\n"]
    if crits or s_crits:
        lines.append("🚨 *Критичные:*")
        lines.extend(crits + s_crits)
    if warns or s_warns:
        lines.append("\n⚠️ *Предупреждения:*")
        lines.extend(warns + s_warns)
    if rolling_notice:
        lines.append(f"\n{rolling_notice}")
    await answer_markdown(msg, "\n".join(lines))


# ── ДУБЛИРОВАНИЕ ВЕСОВ ───────────────────────────────────
@dp.message(F.text == "🔍 Дублирование весов")
async def report_doubles(msg: Message, state: FSMContext):
    await state.clear()
    active = db_get_active_report(msg.from_user.id)
    if not active:
        await msg.answer("📭 Сначала загрузите свежий файл отчёта.")
        return
    rows = db_get_active_month_data(active)
    if not rows:
        await msg.answer("📭 Нет завершённых суток.")
        return

    lines = [
        "🔍 *Расхождение основных и дублирующих весов*",
        "_норма <0.5%, критично >2%_\n",
    ]
    lines.append("`Дн  Кв4 vs 4Д     Кв3 vs 3Д`")
    for r in rows:
        st4, p4, _ = check_doubles(r.get("kv4", 0), r.get("kv4d", 0))
        st3, p3, _ = check_doubles(r.get("kv3", 0), r.get("kv3d", 0))
        lines.append(
            f"`{r['day_num']:>2d}   {em_dup(st4)}{p4:>5.2f}%      {em_dup(st3)}{p3:>5.2f}%`"
        )
    await answer_markdown(msg, "\n".join(lines))


# ── СКЛАД ВЛАЖНОГО КОНЦЕНТРАТА ────────────────────────────
@dp.message(F.text == "🏭 Склад концентрата")
async def report_stock(msg: Message, state: FSMContext):
    await state.clear()
    active = db_get_active_report(msg.from_user.id)
    if not active:
        await msg.answer("📭 Сначала загрузите свежий файл отчёта.")
        return
    year, month = active["year"], active["month"]
    history = db_get_active_stock_history(active)
    if not history:
        await msg.answer(
            "📭 Нет данных по складу за период последнего отчёта.\n"
            "Используйте *📥 Ввести запас* чтобы добавить данные.",
            parse_mode="Markdown",
        )
        return

    lines = [
        f"🏭 *Склад влажного концентрата · {month:02d}/{year}*\n",
        "`Дн  Произв.  Отгруж.  Вес.изм  Мркш.изм  Несовп`",
    ]
    for r in history:
        nesovp = r.get("nesovpadenie", 0)
        em = "🚨" if abs(nesovp) > STOCK_DIFF_WARN else "✅"
        # [FIX]: Убран знак '+', вызывавший ошибку "Sign not allowed in string format specifier"
        lines.append(
            f"`{r['day_num']:>2d}  "
            f"{fmt2(r.get('produced', 0)):>7s}  "
            f"{fmt2(r.get('shipped', 0)):>7s}  "
            f"{fmt2(r.get('ves_izm', 0)):>7s}  "
            f"{fmt2(r.get('marksh_izm', 0)):>7s}  "
            f"{em}{nesovp:+.0f}`"
        )

    last = history[-1]
    lines.append(
        f"\n📅 Последний расчёт: день {last['day_num']}\n"
        f"⚙️ Произведено:  {fmt(last.get('produced', 0))} т\n"
        f"🚂 Отгружено:    {fmt(last.get('shipped', 0))} т\n"
        f"⚖️ Вес.изм.:     {fmt(last.get('ves_izm', 0))} т\n"
        f"📐 Маркш.изм.:   {fmt(last.get('marksh_izm', 0))} т\n"
        f"❗ Несовпадение: {fmt(last.get('nesovpadenie', 0))} т"
    )
    await answer_markdown(msg, "\n".join(lines))


# ── ВВОД ЗАПАСОВ СКЛАДА ──────────────────────────────────
@dp.message(F.text == "📥 Ввести запас")
async def stock_input_start(msg: Message, state: FSMContext):
    await state.clear()
    active = db_get_active_report(msg.from_user.id)
    if not active:
        await msg.answer(
            "📭 Нет выбранного отчёта. Сначала загрузите *report.xls*.",
            parse_mode="Markdown",
        )
        return
    rows = db_get_active_month_data(active)
    if not rows:
        await msg.answer(
            "📭 Нет данных из отчёта. Сначала загрузите *report.xls*.",
            parse_mode="Markdown",
        )
        return

    row = rows[-1]
    day_num = row["day_num"]
    db_year = row["year"]
    db_month = row["month"]

    await state.update_data(day_num=day_num, year=db_year, month=db_month)
    await state.set_state(StockInput.waiting_for_prev)

    await msg.answer(
        f"📥 *Склад влажного концентрата — день "
        f"{day_label(db_year, db_month, day_num)}*\n\n"
        f"Шаг 1 из 2\n"
        f"Введите запас на складе на конец "
        f"*{previous_day_label(db_year, db_month, day_num)}* (тонн):",
        parse_mode="Markdown",
    )


@dp.message(StockInput.waiting_for_prev)
async def stock_input_prev(msg: Message, state: FSMContext):
    # [FIX]: Прерываем ввод, если нажата кнопка меню, чтобы FSM не залипал
    if _is_cancel_command(msg):
        await state.clear()
        await msg.answer("Действие отменено. Выберите команду из меню заново.")
        return

    try:
        stock_prev = parse_user_number(msg.text)
    except ValueError:
        await msg.answer(
            "❌ Введите число (тонн). Например: `61500`", parse_mode="Markdown"
        )
        return

    data = await state.get_data()
    await state.update_data(stock_prev=stock_prev)
    await state.set_state(StockInput.waiting_for_curr)

    day_num = data["day_num"]
    db_year = data["year"]
    db_month = data["month"]
    await msg.answer(
        f"Шаг 2 из 2\n"
        f"Введите запас на складе на конец "
        f"*{day_label(db_year, db_month, day_num)}* (тонн):",
        parse_mode="Markdown",
    )


@dp.message(StockInput.waiting_for_curr)
async def stock_input_curr(msg: Message, state: FSMContext):
    # [FIX]: Защита FSM
    if _is_cancel_command(msg):
        await state.clear()
        await msg.answer("Действие отменено. Выберите команду из меню заново.")
        return

    try:
        stock_curr = parse_user_number(msg.text)
    except ValueError:
        await msg.answer(
            "❌ Введите число (тонн). Например: `52500`", parse_mode="Markdown"
        )
        return

    data = await state.get_data()
    await state.clear()
    year = data["year"]
    month = data["month"]
    day_num = data["day_num"]
    stock_prev = data["stock_prev"]

    try:
        result = db_save_stock(
            year, month, day_num, stock_prev, stock_curr, msg.from_user.id
        )
    except (ReportDataError, sqlite3.Error) as exc:
        logger.error("Ошибка расчёта склада: %s", exc)
        await msg.answer(f"❌ Расчёт не сохранён: {exc}")
        return

    ves_izm = result["ves_izm"]
    marksh_izm = result["marksh_izm"]
    nesovp = result["nesovpadenie"]

    lines = [
        f"✅ *Расчёт склада — день {day_label(year, month, day_num)}*\n",
        f"📦 Запас на {previous_day_label(year, month, day_num)}: {fmt(stock_prev)} т",
        f"📦 Запас на {day_label(year, month, day_num)}: {fmt(stock_curr)} т",
        "",
        f"⚙️ Произведено:   {fmt(result['produced'])} т",
        f"🚂 Отгружено:     {fmt(result['shipped'])} т",
        f"⚖️ Вес.изменение: {fmt(ves_izm)} т",
        f"📐 Маркш.изм.:    {fmt(marksh_izm)} т",
        f"❗ Несовпадение:  {fmt(nesovp)} т",
    ]

    if stock_curr < 0:
        lines.append("\n🚨 *АЛЕРТ: запас отрицательный! Проверьте данные.*")
    elif abs(nesovp) > STOCK_DIFF_WARN:
        em = "🚨" if abs(nesovp) > STOCK_DIFF_WARN * 2 else "⚠️"
        lines.append(
            f"\n{em} *Несовпадение {fmt(abs(nesovp))} т > 500 т!*\n"
            f"Проверьте весы 44/46/74 и 65/66/84."
        )
    else:
        lines.append("\n✅ Несовпадение в норме (< 500 т)")

    await answer_markdown(msg, "\n".join(lines))


# ── СКЛАД — НОЧНАЯ СМЕНА ─────────────────────────────────
@dp.message(F.text == "🌅 Склад — ночная смена")
async def stock_night_start(msg: Message, state: FSMContext):
    await state.clear()
    active = db_get_active_report(msg.from_user.id)
    if not active:
        await msg.answer(
            "📭 Нет выбранного отчёта. Сначала загрузите *report.xls*.",
            parse_mode="Markdown",
        )
        return
    ns = db_get_night_shift(active["year"], active["month"])
    if not ns:
        await msg.answer(
            "📭 Нет данных по ночной смене.\nСначала загрузите *report.xls*.",
            parse_mode="Markdown",
        )
        return

    day_num = ns["day_num"]
    db_year = ns["year"]
    db_month = ns["month"]
    produced = calc_produced(ns)
    shipped = calc_shipped(ns)

    await state.update_data(
        day_num=day_num,
        year=db_year,
        month=db_month,
        produced=produced,
        shipped=shipped,
    )
    await state.set_state(StockNightInput.waiting_for_night)

    await msg.answer(
        f"🌅 *Склад — ночная смена {day_label(db_year, db_month, day_num)}*\n\n"
        f"_Смена 1 (19:30–07:30) из отчёта:_\n"
        f"⚙️ Произведено: {fmt(produced)} т\n"
        f"🚂 Отгружено:   {fmt(shipped)} т\n\n"
        f"Шаг 1 из 2\n"
        f"Введите запас на складе на *начало смены* "
        f"(вечер {previous_day_label(db_year, db_month, day_num)}, тонн):",
        parse_mode="Markdown",
    )


@dp.message(StockNightInput.waiting_for_night)
async def stock_night_prev(msg: Message, state: FSMContext):
    # [FIX]: Защита FSM
    if _is_cancel_command(msg):
        await state.clear()
        await msg.answer("Действие отменено. Выберите команду из меню заново.")
        return

    try:
        stock_night = parse_user_number(msg.text)
    except ValueError:
        await msg.answer(
            "❌ Введите число (тонн). Например: `61500`", parse_mode="Markdown"
        )
        return

    data = await state.get_data()
    await state.update_data(stock_night=stock_night)
    await state.set_state(StockNightInput.waiting_for_morning)

    day_num = data["day_num"]
    year = data["year"]
    month = data["month"]
    await msg.answer(
        f"Шаг 2 из 2\n"
        f"Введите запас на складе на *конец смены* "
        f"(утро {day_label(year, month, day_num)}, тонн):",
        parse_mode="Markdown",
    )


@dp.message(StockNightInput.waiting_for_morning)
async def stock_night_curr(msg: Message, state: FSMContext):
    # [FIX]: Защита FSM
    if _is_cancel_command(msg):
        await state.clear()
        await msg.answer("Действие отменено. Выберите команду из меню заново.")
        return

    try:
        stock_morning = parse_user_number(msg.text)
    except ValueError:
        await msg.answer(
            "❌ Введите число (тонн). Например: `58000`", parse_mode="Markdown"
        )
        return

    data = await state.get_data()
    await state.clear()

    day_num = data["day_num"]
    year = data["year"]
    month = data["month"]
    produced = data["produced"]
    shipped = data["shipped"]
    stock_night = data["stock_night"]

    ves_izm = produced - shipped
    marksh_izm = stock_morning - stock_night
    nesovp = ves_izm - marksh_izm

    lines = [
        f"🌅 *Склад — ночная смена {day_label(year, month, day_num)}*\n",
        f"📦 Запас на начало смены: {fmt(stock_night)} т",
        f"📦 Запас на конец смены:  {fmt(stock_morning)} т",
        "",
        f"⚙️ Произведено:   {fmt(produced)} т",
        f"🚂 Отгружено:     {fmt(shipped)} т",
        f"⚖️ Вес.изменение: {fmt(ves_izm)} т",
        f"📐 Маркш.изм.:    {fmt(marksh_izm)} т",
        f"❗ Несовпадение:  {fmt(nesovp)} т",
    ]

    if stock_morning < 0:
        lines.append("\n🚨 *АЛЕРТ: запас отрицательный! Проверьте данные.*")
    elif abs(nesovp) > STOCK_DIFF_WARN:
        em = "🚨" if abs(nesovp) > STOCK_DIFF_WARN * 2 else "⚠️"
        lines.append(
            f"\n{em} *Несовпадение {fmt(abs(nesovp))} т > 500 т!*\n"
            f"Проверьте весы 44/46/74 и 65/66/84."
        )
    else:
        lines.append("\n✅ Несовпадение в норме (< 500 т)")

    await answer_markdown(msg, "\n".join(lines))


# ── AI ────────────────────────────────────────────────────
@dp.message(F.text == "🤖 Спросить AI")
async def ai_request(msg: Message, state: FSMContext):
    await state.clear()
    await state.set_state(AIState.waiting_for_question)
    await msg.answer(
        "🤖 *AI-Агент метролога ДОФ*\nВведите вопрос по балансу, нормам или диагностике весов:",
        parse_mode="Markdown",
    )


@dp.message(AIState.waiting_for_question)
async def ai_processing(msg: Message, state: FSMContext):
    # [FIX]: Защита FSM
    if _is_cancel_command(msg):
        await state.clear()
        await msg.answer("Диалог с AI отменен. Выберите команду меню.")
        return

    await state.clear()
    wait = await msg.answer("⏳ Анализирую...")
    active = db_get_active_report(msg.from_user.id)
    rows = db_get_active_month_data(active)
    end_date = active_report_end_date(active)
    rolling_rows = (
        db_get_latest_completed_days(3, end_date=end_date)
        if end_date is not None
        else []
    )
    context = make_ai_context(rows, rolling_rows) if rows else "Нет данных."
    answer = await ask_ai(msg.text, context)
    chunks = split_message(f"🤖 AI-Агент:\n\n{answer}")
    await wait.edit_text(chunks[0])
    for chunk in chunks[1:]:
        await msg.answer(chunk)


# ── ПОМОЩЬ ────────────────────────────────────────────────
@dp.message(F.text == "❓ Помощь")
async def h_help(msg: Message, state: FSMContext):
    await state.clear()
    await msg.answer(
        "📖 *Как пользоваться*\n\n"
        "1️⃣ Отправьте файл `report.xls` боту\n"
        "2️⃣ Бот прочитает все данные автоматически\n"
        "3️⃣ Текущий незавершённый день полностью исключается из суточного и "
        "скользящего баланса; архивные завершённые дни сохраняются полностью\n\n"
        "*Кнопки:*\n"
        "📅 Суточный баланс — таблица по дням (только полные сутки)\n"
        "🌙 Ночная смена — Смена 1 текущего незавершённого дня, отдельным балансом\n"
        "📉 Просмотр проскальзывания — скользящий баланс за 1, 2 и 3 завершённых суток\n"
        "📆 Недельная сводка — за последние 7 полных суток, включая границу месяца\n"
        "🗓 Месячный итог — накопительный баланс (без текущего дня)\n"
        "🔔 Алерты — суточные нормы и дубли; сигнал баланса только по 3 суткам\n"
        "🔍 Дублирование весов — Конв.4/4Д и Конв.3/3Д\n"
        "🤖 Спросить AI — диагностика и рекомендации\n\n"
        "*Нормы конвейеров:*\n"
        "Конв.14/15: 60–80%  |  Конв.101/102: 6–20%\n"
        "Конв.19/31/32: 40–60%  |  Конв.33/34: 83–87%\n"
        "Конв.10/34А: 15–45%\n\n"
        "*Баланс:* ✅ ±2%  ⚠️ ±5%  🚨 >5%\n"
        "*Скользящий расчёт:* сначала суммируется тоннаж, затем вычисляется процент\n"
        "*Дублирование:* ✅ <0.5%  ⚠️ <2%  🚨 >2%",
        parse_mode="Markdown",
    )


# ════════════════════════════════════════════════════════
#  ЗАПУСК
# ════════════════════════════════════════════════════════
async def main():
    init_db()
    logger.info("ДОФ Баланс Bot запущен.")
    await dp.start_polling(bot)


if __name__ == "__main__":
    import sys

    print("=== СТАРТ ПРИЛОЖЕНИЯ ===")
    try:
        print(f"BOT_TOKEN установлен: {bool(BOT_TOKEN)}")
        asyncio.run(main())
    except Exception as e:
        print(f"ОШИБКА ПРИ ЗАПУСКЕ: {e}", file=sys.stderr)
        raise
