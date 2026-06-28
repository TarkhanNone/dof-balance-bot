import asyncio
import sqlite3
import os
import io
import logging
from datetime import datetime
from typing import Optional

from dotenv import load_dotenv
import xlrd
import openpyxl

import aiohttp
from aiogram import Bot, Dispatcher, F
from aiogram.filters import Command
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.fsm.storage.memory import MemoryStorage
from aiogram.types import (
    Message, Document,
    ReplyKeyboardMarkup, KeyboardButton,
)

load_dotenv()
logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
logger = logging.getLogger(__name__)

# ════════════════════════════════════════════════════════
#  НАСТРОЙКИ
# ════════════════════════════════════════════════════════
BOT_TOKEN         = os.getenv("BOT_TOKEN")
ANTHROPIC_API_KEY = os.getenv("ANTHROPIC_API_KEY")
DB_PATH           = "dof_balance.db"
AI_MODEL          = "claude-sonnet-4-6"   # актуальная модель

if not BOT_TOKEN:
    raise RuntimeError("BOT_TOKEN не задан (переменная окружения)")

# ════════════════════════════════════════════════════════
#  НОРМЫ БАЛАНСА (% от Конв.3 или Конв.4 = 100%)
#  ключ: (мин%, макс%, человекочитаемое имя, очередь)
# ════════════════════════════════════════════════════════
NORMS = {
    "kv14":  (60, 80, "Конв.14 (дробление/грохочение)", 1),
    "kv15":  (60, 80, "Конв.15 (дробление/грохочение)", 2),
    "kv32":  (40, 60, "Конв.32 (ПП после сепарации)",   1),
    "kv31":  (40, 60, "Конв.31 (ПП после сепарации)",   2),
    "kv34":  (83, 87, "Конв.34 (ПП → ММС)",             1),
    "kv33":  (83, 87, "Конв.33 (ПП → ММС)",             2),
    "kv102": (6,  20, "Конв.102 (хвосты → склад 105)",  1),
    "kv101": (6,  20, "Конв.101 (хвосты → склад 105)",  2),
    "kv19":  (40, 60, "Конв.19 (ПП 2 очереди)",         2),
    "kv34a": (15, 45, "Конв.10/34А (мелкая фракция)",   1),
}
WARN_PCT = 2.0   # баланс: предупреждение
CRIT_PCT = 5.0   # баланс: критично
DUP_WARN = 0.5   # дублирующие весы: предупреждение, %
DUP_CRIT = 2.0   # дублирующие весы: критично, %

# Маппинг названий конвейеров из отчёта → ключи БД
CONV_MAP = {
    "Конвейер 4":    "kv4",   "Конвейер 4Д":   "kv4d",
    "Конвейер 14":   "kv14",  "Конвейер 32":   "kv32",
    "Конвейер 18":   "kv18",  "Конвейер 102":  "kv102",
    "Конвейер 34":   "kv34",
    "Конвейер 24ПП": "kv24p", "Конвейер 24ХВ": "kv24hv",
    "Конвейер 28А.1":"kv28a1","Конвейер 34A":  "kv34a", "Конвейер 34А": "kv34a",
    "Конвейер 3":    "kv3",   "Конвейер 3Д":   "kv3d",
    "Конвейер 15":   "kv15",  "Конвейер 31":   "kv31",
    "Конвейер 19":   "kv19",  "Конвейер 101":  "kv101",
    "Конвейер 33":   "kv33",  "Конвейер 28А.2":"kv28a2",
    "Конвейер 44":   "kv44",  "Конвейер 44Д":  "kv44d",
    "Конвейер 46":   "kv46",  "Конвейер 46Д":  "kv46d",
    "Конвейер 74":   "kv74",  "Конвейер 74Д":  "kv74d",
    "Конвейер 65МПС":"kv65mps","Конвейер 65ЦПО":"kv65cpo",
    "Конвейер 66МПС":"kv66mps","Конвейер 66ЦПО":"kv66cpo",
    "Конвейер 84МПС":"kv84mps","Конвейер 84ЦПО":"kv84cpo",
    "Конвейер 63":   "kv63",  "Конвейер 61":   "kv61",
}

FIELDS = [
    "kv4","kv4d","kv14","kv32","kv34","kv34a","kv102","kv24p",
    "kv24hv","kv28a1","kv3","kv3d","kv15","kv19","kv31","kv33",
    "kv101","kv28a2","kv44","kv44d","kv46","kv46d","kv74","kv74d",
    "kv65mps","kv65cpo","kv66mps","kv66cpo","kv84mps","kv84cpo",
    "kv63","kv61",
]

# ════════════════════════════════════════════════════════
#  ФОРМУЛЫ СКЛАДА ВЛАЖНОГО КОНЦЕНТРАТА
# ════════════════════════════════════════════════════════
STOCK_DIFF_WARN = 500   # тонн — порог алерта по расхождению

def calc_produced(d: dict) -> float:
    """Произведено = (44+44Д)/2 + 46Д + (74+74Д)/2"""
    avg44 = (d.get("kv44", 0) + d.get("kv44d", 0)) / 2
    k46d  = d.get("kv46d", 0)
    avg74 = (d.get("kv74", 0) + d.get("kv74d", 0)) / 2
    return avg44 + k46d + avg74

def calc_shipped(d: dict) -> float:
    """Отгружено = 65МПС+65ЦПО+66МПС+66ЦПО+84МПС+84ЦПО"""
    return (d.get("kv65mps", 0) + d.get("kv65cpo", 0) +
            d.get("kv66mps", 0) + d.get("kv66cpo", 0) +
            d.get("kv84mps", 0) + d.get("kv84cpo", 0))

# ════════════════════════════════════════════════════════
#  БАЗА ДАННЫХ (безопасные context manager'ы)
# ════════════════════════════════════════════════════════
def init_db():
    with sqlite3.connect(DB_PATH) as conn:
        cols_sql = ", ".join(f"{f} REAL DEFAULT 0" for f in FIELDS)
        conn.execute(f"""
            CREATE TABLE IF NOT EXISTS daily_data (
                id          INTEGER PRIMARY KEY AUTOINCREMENT,
                report_date TEXT,
                day_num     INTEGER,
                year        INTEGER,
                month       INTEGER,
                {cols_sql},
                source      TEXT,
                uploaded    TEXT DEFAULT CURRENT_TIMESTAMP,
                UNIQUE(year, month, day_num)
            )
        """)
        # Смена 1 (ночная) последнего, ещё незавершённого дня — отдельно
        # от daily_data. НИКОГДА не участвует в суточном/недельном/
        # месячном балансе, только показывается по отдельной кнопке.
        conn.execute(f"""
            CREATE TABLE IF NOT EXISTS night_shift (
                id          INTEGER PRIMARY KEY AUTOINCREMENT,
                year        INTEGER,
                month       INTEGER,
                day_num     INTEGER,
                {cols_sql},
                source      TEXT,
                uploaded    TEXT DEFAULT CURRENT_TIMESTAMP,
                UNIQUE(year, month, day_num)
            )
        """)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS report_log (
                id         INTEGER PRIMARY KEY AUTOINCREMENT,
                filename   TEXT, period TEXT, rows_saved INTEGER,
                user_id    INTEGER, uploaded TEXT DEFAULT CURRENT_TIMESTAMP
            )
        """)
        # Склад влажного концентрата — суточный расчёт.
        # Оператор вводит запас за предыдущий и текущий день.
        # Несовпадение = Вес.изм.сут − Изм.запаса.маркш
        conn.execute("""
            CREATE TABLE IF NOT EXISTS stock_data (
                id              INTEGER PRIMARY KEY AUTOINCREMENT,
                year            INTEGER,
                month           INTEGER,
                day_num         INTEGER,    -- день расчёта (напр. 18)
                stock_prev      REAL,       -- запас на конец дня N-1 (от оператора, напр. за 17-е)
                stock_curr      REAL,       -- запас на конец дня N   (от оператора, напр. за 18-е)
                produced        REAL,       -- произведено за день N (из report.xls)
                shipped         REAL,       -- отгружено за день N (из report.xls)
                ves_izm         REAL,       -- весовое изм. = produced − shipped
                marksh_izm      REAL,       -- маркш. изм. = stock_curr − stock_prev
                nesovpadenie    REAL,       -- несовпадение = ves_izm − marksh_izm
                user_id         INTEGER,
                entered_at      TEXT DEFAULT CURRENT_TIMESTAMP,
                UNIQUE(year, month, day_num)
            )
        """)
        conn.commit()


def db_save_daily(parsed: dict, user_id: int, filename: str, year: int, month: int) -> tuple:
    """
    Сохраняет суточные данные (сумма смена1+смена2 на конвейер).

    ПРАВИЛО ПОЛНОТЫ СУТОК (см. структуру смен ДОФ):
    Сутки попадают в daily_data (и далее в суточный/недельный/месячный
    баланс) ТОЛЬКО когда в файле присутствуют ОБЕ смены этого дня.
    Файл скачивается утром, когда Смена 1 последнего дня уже завершена,
    а Смена 2 ещё идёт — поэтому ПОСЛЕДНИЙ (максимальный) день в файле
    полностью исключается из daily_data, включая его Смену 1.
    Смена 1 последнего дня сохраняется отдельно в night_shift —
    показывается только по отдельной кнопке "🌙 Ночная смена",
    в общий баланс не подмешивается.
    """
    base_fields = ["year", "month", "day_num", "report_date", "source"] + FIELDS
    cols = ",".join(base_fields)
    qs   = ",".join(["?"] * len(base_fields))
    upd  = ",".join(f"{f}=excluded.{f}" for f in FIELDS) + ", source=excluded.source"
    sql_query = (
        f"INSERT INTO daily_data ({cols}) VALUES ({qs}) "
        f"ON CONFLICT(year,month,day_num) DO UPDATE SET {upd}"
    )

    ns_base_fields = ["year", "month", "day_num", "source"] + FIELDS
    ns_cols = ",".join(ns_base_fields)
    ns_qs   = ",".join(["?"] * len(ns_base_fields))
    ns_upd  = ",".join(f"{f}=excluded.{f}" for f in FIELDS) + ", source=excluded.source"
    ns_sql_query = (
        f"INSERT INTO night_shift ({ns_cols}) VALUES ({ns_qs}) "
        f"ON CONFLICT(year,month,day_num) DO UPDATE SET {ns_upd}"
    )

    by_day = parsed.get("daily_by_shift", {})
    all_days = [d for d, v in by_day.items() if v.get(1) or v.get(2)]
    if not all_days:
        return 0, 0
    max_day = max(all_days)

    saved = 0
    with sqlite3.connect(DB_PATH) as conn:
        # Прошлая "ночная смена" этого месяца устарела — сбрасываем
        conn.execute("DELETE FROM night_shift WHERE year=? AND month=?", (year, month))
        # На случай, если последний день раньше уже считался полным
        # (например, файл загрузили повторно в тот же день) — убираем
        # его из daily_data, раз сейчас он снова неполный.
        conn.execute(
            "DELETE FROM daily_data WHERE year=? AND month=? AND day_num=?",
            (year, month, max_day)
        )

        for day_num, shifts in by_day.items():
            s1 = shifts.get(1, {})
            s2 = shifts.get(2, {})
            if not s1 and not s2:
                continue

            if day_num == max_day:
                # Последний день: НЕ идёт в daily_data вообще.
                # Смена 1 сохраняется отдельно для кнопки "Ночная смена".
                if s1 and any(v for v in s1.values()):
                    ns_rec = {"year": year, "month": month, "day_num": day_num, "source": filename}
                    for f in FIELDS:
                        ns_rec[f] = s1.get(f, 0.0)
                    try:
                        conn.execute(ns_sql_query, list(ns_rec.values()))
                    except Exception as e:
                        logger.warning(f"Ошибка сохранения ночной смены дня {day_num}: {e}")
                continue

            # Прошлые (завершённые) дни: обе смены суммируются
            vals = {f: s1.get(f, 0.0) + s2.get(f, 0.0) for f in FIELDS}
            rec = {
                "year": year, "month": month, "day_num": day_num,
                "report_date": f"{year:04d}-{month:02d}-{day_num:02d}",
                "source": filename,
            }
            for f in FIELDS:
                rec[f] = vals.get(f, 0.0)
            try:
                conn.execute(sql_query, list(rec.values()))
                saved += 1
            except Exception as e:
                logger.warning(f"Ошибка сохранения дня {day_num}: {e}")

        conn.execute(
            "INSERT INTO report_log (filename,period,rows_saved,user_id) VALUES (?,?,?,?)",
            (filename, parsed.get("period", ""), saved, user_id)
        )
        conn.commit()
    return saved, max_day


def db_get_night_shift(year: int, month: int) -> Optional[dict]:
    """Смена 1 (ночная) последнего незавершённого дня — для кнопки '🌙 Ночная смена'."""
    with sqlite3.connect(DB_PATH) as conn:
        conn.row_factory = sqlite3.Row
        row = conn.execute(
            "SELECT * FROM night_shift WHERE year=? AND month=? ORDER BY day_num DESC LIMIT 1",
            (year, month)
        ).fetchone()
    return dict(row) if row else None


# ════════════════════════════════════════════════════════
#  СКЛАД ВЛАЖНОГО КОНЦЕНТРАТА — DB ФУНКЦИИ
#
#  ФОРМУЛЫ (восстановлены из Учет_по_конвейерам_2025.xlsx):
#
#  Произведено_нак   = Σ[(44+44Д)/2 + 46Д + (74+74Д)/2]  с 1 по N день
#  Отгружено_нак     = Σ[65МПС+65ЦПО+66МПС+66ЦПО+84МПС+84ЦПО] с 1 по N день
#  СКЛАД ВЛАЖНОГО КОНЦЕНТРАТА — DB ФУНКЦИИ
#
#  Суточный расчёт за день N:
#    Произведено(N)   = (44+44Д)/2 + 46Д + (74+74Д)/2  из report.xls за день N
#    Отгружено(N)     = 65МПС+ЦПО + 66МПС+ЦПО + 84МПС+ЦПО из report.xls за день N
#    Вес.изм.(N)      = Произведено − Отгружено
#    Маркш.изм.(N)    = Запас_N − Запас_(N-1)   (оба числа от оператора)
#    Несовпадение(N)  = Вес.изм. − Маркш.изм.   (алерт если |Несовп| > 500 т)
# ════════════════════════════════════════════════════════
def db_save_stock(year: int, month: int, day_num: int,
                  stock_prev: float, stock_curr: float,
                  user_id: int) -> dict:
    """
    Сохраняет суточный расчёт склада за день day_num.
    stock_prev — запас на конец дня (N-1), stock_curr — на конец дня N.
    Данные Произведено/Отгружено берутся из daily_data автоматически.
    """
    with sqlite3.connect(DB_PATH) as conn:
        conn.row_factory = sqlite3.Row
        row = conn.execute(
            "SELECT * FROM daily_data WHERE year=? AND month=? AND day_num=?",
            (year, month, day_num)
        ).fetchone()

    if row:
        d = dict(row)
        produced = calc_produced(d)
        shipped  = calc_shipped(d)
    else:
        produced = 0.0
        shipped  = 0.0

    ves_izm      = produced - shipped
    marksh_izm   = stock_curr - stock_prev
    nesovpadenie = ves_izm - marksh_izm

    with sqlite3.connect(DB_PATH) as conn:
        conn.execute("""
            INSERT INTO stock_data
                (year, month, day_num, stock_prev, stock_curr,
                 produced, shipped, ves_izm, marksh_izm, nesovpadenie, user_id)
            VALUES (?,?,?,?,?,?,?,?,?,?,?)
            ON CONFLICT(year,month,day_num) DO UPDATE SET
                stock_prev=excluded.stock_prev,
                stock_curr=excluded.stock_curr,
                produced=excluded.produced,
                shipped=excluded.shipped,
                ves_izm=excluded.ves_izm,
                marksh_izm=excluded.marksh_izm,
                nesovpadenie=excluded.nesovpadenie,
                user_id=excluded.user_id,
                entered_at=CURRENT_TIMESTAMP
        """, (year, month, day_num, stock_prev, stock_curr,
              produced, shipped, ves_izm, marksh_izm, nesovpadenie, user_id))
        conn.commit()

    return {
        "day_num":      day_num,
        "produced":     produced,
        "shipped":      shipped,
        "ves_izm":      ves_izm,
        "marksh_izm":   marksh_izm,
        "nesovpadenie": nesovpadenie,
    }


def db_get_stock_history(year: int, month: int) -> list:
    """История склада за месяц."""
    with sqlite3.connect(DB_PATH) as conn:
        conn.row_factory = sqlite3.Row
        rows = conn.execute(
            "SELECT * FROM stock_data WHERE year=? AND month=? ORDER BY day_num ASC",
            (year, month)
        ).fetchall()
    return [dict(r) for r in rows]


def db_get_month_data(year: int, month: int) -> list:
    with sqlite3.connect(DB_PATH) as conn:
        conn.row_factory = sqlite3.Row
        rows = conn.execute(
            "SELECT * FROM daily_data WHERE year=? AND month=? ORDER BY day_num ASC",
            (year, month)
        ).fetchall()
    return [dict(r) for r in rows]


# ════════════════════════════════════════════════════════
#  ПАРСЕР EXCEL — без дублирования (DRY), .append(), try/except
# ════════════════════════════════════════════════════════
def _safe(v) -> float:
    try:
        return float(v) if v is not None else 0.0
    except (ValueError, TypeError):
        return 0.0


def _pick_sheet(sheet_names: list, preferred: list, fallback_idx: int = 0) -> str:
    for sh in preferred:
        if sh in sheet_names:
            return sh
    return sheet_names[fallback_idx]


def _read_xls_sheet(file_bytes: bytes, sheet_name: str) -> list:
    wb = xlrd.open_workbook(file_contents=file_bytes)
    if sheet_name not in wb.sheet_names():
        return []
    ws = wb.sheet_by_name(sheet_name)
    rows = []
    for r in range(ws.nrows):
        rows.append([ws.cell_value(r, c) for c in range(ws.ncols)])
    return rows


def _read_xlsx_sheet(file_bytes: bytes, sheet_name: str) -> list:
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    return list(ws.iter_rows(values_only=True))


def _get_sheet_names(file_bytes: bytes, filename: str) -> list:
    if filename.lower().endswith(".xls"):
        wb = xlrd.open_workbook(file_contents=file_bytes)
        return wb.sheet_names()
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    return wb.sheetnames


def parse_report(file_bytes: bytes, filename: str) -> dict:
    """
    Парсит лист 'ДетСменТекМесяц' — детализация по сменам (Смена 1 / Смена 2)
    для каждого дня текущего месяца.

    Структура листа (см. фото отчёта пользователя):
      строка с "Дата:"  — номер дня, указан НАД парой колонок (Смена1, Смена2),
                           т.е. день стоит только в нечётной колонке пары
      строка с "Смена:" — 1 или 2 для каждой колонки
      далее построчно   — название конвейера, затем значения по всем колонкам

    Возвращает:
      daily_by_shift: { day_num: { 1: {kv4:.., ...}, 2: {kv4:.., ...} } }
      period, used_sheet, error
    """
    result = {"period": "", "daily_by_shift": {}, "sheets_found": [],
              "used_sheet": "", "error": ""}

    try:
        sheet_names = _get_sheet_names(file_bytes, filename)
    except Exception as err:
        logger.error(f"Ошибка чтения файла {filename}: {err}")
        result["error"] = f"Не удалось открыть файл как Excel-таблицу: {err}"
        return result

    result["sheets_found"] = sheet_names

    sheet_name = "ДетСменТекМесяц"
    if sheet_name not in sheet_names:
        result["error"] = (
            f"В файле не найден лист '{sheet_name}'. "
            f"Доступные листы: {', '.join(sheet_names)}"
        )
        return result
    result["used_sheet"] = sheet_name

    try:
        if filename.lower().endswith(".xls"):
            rows = _read_xls_sheet(file_bytes, sheet_name)
        else:
            rows = _read_xlsx_sheet(file_bytes, sheet_name)
    except Exception as err:
        logger.error(f"Ошибка чтения листа {sheet_name}: {err}")
        result["error"] = f"Не удалось прочитать лист '{sheet_name}': {err}"
        return result

    # Заголовок периода — первая строка с текстом вида "...c 1 по 19"
    if rows and rows[0] and rows[0][0]:
        result["period"] = str(rows[0][0]).strip()

    # Найти строку "Дата:" и строку "Смена:"
    date_row_idx = None
    for i, row in enumerate(rows):
        if row and str(row[0]).strip() in ("Дата:", "Дата"):
            date_row_idx = i
            break

    if date_row_idx is None:
        result["error"] = f"На листе '{sheet_name}' не найдена строка 'Дата:'."
        return result

    shift_row_idx = date_row_idx + 1
    if shift_row_idx >= len(rows) or str(rows[shift_row_idx][0]).strip() not in ("Смена:", "Смена"):
        result["error"] = f"На листе '{sheet_name}' не найдена строка 'Смена:' сразу после 'Дата:'."
        return result

    date_row = rows[date_row_idx]
    shift_row = rows[shift_row_idx]

    # Карта колонок: col_idx -> (day_num, shift_num)
    # День проставлен только в первой колонке пары, поэтому "тащим" значение вперёд
    col_map = {}
    current_day = None
    for col_i in range(1, max(len(date_row), len(shift_row))):
        d = date_row[col_i] if col_i < len(date_row) else None
        s = shift_row[col_i] if col_i < len(shift_row) else None
        try:
            if d is not None and str(d).strip() != "":
                current_day = int(float(d))
        except (ValueError, TypeError):
            pass
        try:
            shift_num = int(float(s)) if s is not None and str(s).strip() != "" else None
        except (ValueError, TypeError):
            shift_num = None
        if current_day is not None and shift_num in (1, 2):
            col_map[col_i] = (current_day, shift_num)

    if not col_map:
        result["error"] = "Не удалось сопоставить колонки с днями и сменами."
        return result

    # Данные конвейеров построчно
    daily_by_shift = {}
    for row in rows[shift_row_idx + 1:]:
        if not row or not row[0]:
            continue
        name = str(row[0]).strip()
        key = CONV_MAP.get(name) or CONV_MAP.get(name.rstrip())
        if not key:
            continue
        for col_i, (day_num, shift_num) in col_map.items():
            if col_i >= len(row):
                continue
            val = _safe(row[col_i])
            daily_by_shift.setdefault(day_num, {1: {}, 2: {}})
            daily_by_shift[day_num][shift_num][key] = val

    result["daily_by_shift"] = daily_by_shift
    return result



# ════════════════════════════════════════════════════════
#  ФОРМУЛЫ БАЛАНСА (из Баланс_ДО_2026-1.xlsx, проверено)
# ════════════════════════════════════════════════════════
def bal1(d: dict) -> Optional[float]:
    """Баланс 1 = 102+34+24П+24Хв−28А.I−4, % от Конв.4"""
    b = d.get("kv4", 0)
    if not b:
        return None
    return (d.get("kv102", 0) + d.get("kv34", 0) + d.get("kv24p", 0)
            + d.get("kv24hv", 0) - d.get("kv28a1", 0) - b) / b * 100


def bal2(d: dict) -> Optional[float]:
    """Баланс 2 = 101+33−28А.II−3, % от Конв.3"""
    b = d.get("kv3", 0)
    if not b:
        return None
    return (d.get("kv101", 0) + d.get("kv33", 0) - d.get("kv28a2", 0) - b) / b * 100


def balc1(d: dict) -> Optional[float]:
    """Баланс С.1 = 102+24Хв+24П+32−28А.I−14"""
    b = d.get("kv4", 0)
    if not b:
        return None
    return (d.get("kv102", 0) + d.get("kv24hv", 0) + d.get("kv24p", 0)
            + d.get("kv32", 0) - d.get("kv28a1", 0) - d.get("kv14", 0)) / b * 100


def balc2(d: dict) -> Optional[float]:
    """Баланс С.2 = 101+31−28А.II−15"""
    b = d.get("kv3", 0)
    if not b:
        return None
    return (d.get("kv101", 0) + d.get("kv31", 0)
            - d.get("kv28a2", 0) - d.get("kv15", 0)) / b * 100


def pct(v: float, base: float) -> float:
    return v / base * 100 if base else 0.0


def check_norm(val: float, base: float, key: str) -> tuple:
    if not base or key not in NORMS:
        return "none", 0.0
    p = pct(val, base)
    mn, mx = NORMS[key][0], NORMS[key][1]
    margin = (mx - mn) * 0.5
    if mn <= p <= mx:
        return "ok", p
    if mn - margin <= p <= mx + margin:
        return "warn", p
    return "crit", p


def check_doubles(val_main: float, val_dup: float) -> tuple:
    """Сравнение основных и дублирующих весов. Возвращает (статус, разница%, разница_тонн)."""
    if not val_main or not val_dup:
        return "none", 0.0, 0.0
    diff_t = val_main - val_dup
    diff_p = abs(diff_t) / val_main * 100
    if diff_p <= DUP_WARN:
        return "ok", diff_p, diff_t
    if diff_p <= DUP_CRIT:
        return "warn", diff_p, diff_t
    return "crit", diff_p, diff_t


# ════════════════════════════════════════════════════════
#  ФОРМАТИРОВАНИЕ
# ════════════════════════════════════════════════════════
def em_bal(v: Optional[float]) -> str:
    if v is None:
        return "⬜"
    a = abs(v)
    return "✅" if a <= WARN_PCT else "⚠️" if a <= CRIT_PCT else "🚨"


def em_norm(st: str) -> str:
    return {"ok": "✅", "warn": "⚠️", "crit": "🚨", "none": "⬜"}.get(st, "⬜")


def em_dup(st: str) -> str:
    return {"ok": "✅", "warn": "⚠️", "crit": "🚨", "none": "⬜"}.get(st, "⬜")


def sign(v: Optional[float], d: int = 2) -> str:
    if v is None:
        return "—"
    return f"+{v:.{d}f}%" if v >= 0 else f"{v:.{d}f}%"


def fmt(v) -> str:
    if not v and v != 0:
        return "—"
    return f"{int(v):,}".replace(",", " ")


def fmt2(v) -> str:
    if not v and v != 0:
        return "—"
    return f"{v/1000:.1f}k" if abs(v) >= 1000 else f"{v:.0f}"


def build_alerts(d: dict, label: str = "") -> list:
    """Возвращает список (level, text) — все нарушения по одной записи."""
    alerts = []
    base4, base3 = d.get("kv4", 0), d.get("kv3", 0)
    prefix = f"[{label}] " if label else ""

    for name, val in [("Баланс 1", bal1(d)), ("Баланс 2", bal2(d)),
                       ("Баланс С.1", balc1(d)), ("Баланс С.2", balc2(d))]:
        if val is None:
            continue
        if abs(val) > CRIT_PCT:
            alerts.append(("crit", f"🚨 {prefix}{name} = {sign(val)} (критично, норма ±{CRIT_PCT:.0f}%)"))
        elif abs(val) > WARN_PCT:
            alerts.append(("warn", f"⚠️ {prefix}{name} = {sign(val)} (норма ±{WARN_PCT:.0f}%)"))

    checks = [("kv14", base4), ("kv32", base4), ("kv34", base4), ("kv34a", base4),
              ("kv102", base4), ("kv15", base3), ("kv19", base3), ("kv31", base3),
              ("kv33", base3), ("kv101", base3)]
    for key, base in checks:
        val = d.get(key, 0)
        if not val or not base:
            continue
        st, p = check_norm(val, base, key)
        if st in ("warn", "crit"):
            mn, mx, desc, _ = NORMS[key]
            alerts.append((st, f"{em_norm(st)} {prefix}{desc}: {p:.1f}% (норма {mn}–{mx}%)"))

    st4, p4, t4 = check_doubles(base4, d.get("kv4d", 0))
    if st4 in ("warn", "crit"):
        alerts.append((st4, f"{em_dup(st4)} {prefix}Конв.4 vs 4Д: расхождение {p4:.2f}% ({fmt(t4)} т)"))
    st3, p3, t3 = check_doubles(base3, d.get("kv3d", 0))
    if st3 in ("warn", "crit"):
        alerts.append((st3, f"{em_dup(st3)} {prefix}Конв.3 vs 3Д: расхождение {p3:.2f}% ({fmt(t3)} т)"))

    return alerts


# ════════════════════════════════════════════════════════
#  AI АГЕНТ — полный технологический контекст
# ════════════════════════════════════════════════════════
SYSTEM_PROMPT = """Ты — AI-агент метролога горно-обогатительной фабрики (ДОФ), Костанайский регион, Казахстан.

ТЕХНОЛОГИЧЕСКАЯ СХЕМА:
Очередь 1: Руда → Конв.4 (6400 т/ч) → 6 бункеров (6000т каждый) → дробление/грохочение
→ Конв.14 (2500 т/ч) → 5 бункеров (1000т) → Сепарация →
  [ПП]: Конв.18(условный)→Конв.32(2500 т/ч)→Конв.34(3200 т/ч)→ММС
  [Хвосты]: Конв.20(условный)→Конв.24(2500 т/ч)→Склад хв.№25  ИЛИ  Конв.102(800 т/ч)→Склад хв.№105
  [Мелкая фракция]: Конв.10/34А(1250 т/ч)→Конв.34
  [Склад ПП]: Конв.24→Склад ПП; Склад ПП→Конв.28А.I→Конв.32→Конв.34
НЕ в балансе: Конв.18, Конв.20

Очередь 2: Руда → Конв.3 (6400 т/ч) → 6 бункеров (6000т) → дробление/грохочение
→ Конв.15 (2500 т/ч) → 5 бункеров (1000т) → Сепарация →
  [ПП]: Конв.19(2500 т/ч)→Конв.31→Конв.33→ММС
  [Хвосты]: Конв.101(800 т/ч)→Склад хв.№105
  [Склад ПП]: Конв.28А.II→Конв.31→Конв.33

НОРМЫ БАЛАНСА (% от Конв.3 или Конв.4 = 100%):
• Конв.14/15: 70% ±10% (60–80%) — задержка руды в бункерах объясняет расхождение, это НОРМАЛЬНО
• Конв.101/102: 13% ±7% (6–20%) — хвосты сепарации
• Конв.19/31/32: 50% ±10% (40–60%) — промпродукт
• Конв.33/34: 83–87% — промпродукт + мелкая фракция + склад ПП → ММС
• Конв.10/34А: 30% ±15% (15–45%) — мелкая фракция после грохочения

ФОРМУЛЫ БАЛАНСОВ:
Баланс1  = 102+34+24П+24Хв−28А.I−4   (% от Конв.4)
БалансС1 = 102+24Хв+24П+32−28А.I−14
Баланс2  = 101+33−28А.II−3            (% от Конв.3)
БалансС2 = 101+31−28А.II−15

ПОРОГИ: ±2% предупреждение, ±5% критично
Дублирующие весы (Конв.4/4Д, Конв.3/3Д): норма расхождения <0.5%, критично >2%

Объясняй причины отклонений с учётом технологии. Давай конкретные метрологические
рекомендации (поверка, чистка датчиков, проверка калибровки, осмотр ленты).
Отвечай по-русски, кратко, для Telegram. Используй ✅⚠️🚨🔧📊🔮."""


async def ask_ai(question: str, context: str) -> str:
    if not ANTHROPIC_API_KEY:
        return "⚠️ AI недоступен: не задан ANTHROPIC_API_KEY на сервере."

    payload = {
        "model": AI_MODEL,
        "max_tokens": 1000,
        "system": SYSTEM_PROMPT,
        "messages": [{"role": "user", "content": f"{question}\n\nДанные:\n{context}"}],
    }
    try:
        async with aiohttp.ClientSession() as session:
            async with session.post(
                "https://api.anthropic.com/v1/messages",
                headers={
                    "x-api-key": ANTHROPIC_API_KEY,
                    "anthropic-version": "2023-06-01",
                    "content-type": "application/json",
                },
                json=payload,
            ) as resp:
                if resp.status != 200:
                    body = await resp.text()
                    logger.error(f"Anthropic API error {resp.status}: {body}")
                    return f"⚠️ Ошибка AI-сервера (код {resp.status})"
                data = await resp.json()
                return data["content"][0]["text"]
    except Exception as e:
        logger.error(f"AI connection error: {e}")
        return f"⚠️ Не удалось связаться с AI: {e}"


def make_ai_context(rows: list) -> str:
    """Контекст для AI: сводка + список нарушений по полным суткам."""
    full_days = rows  # все записи в daily_data уже полные сутки
    if not full_days:
        return "Нет завершённых суток с данными."

    s = {k: sum(r.get(k, 0) for r in full_days) for k in FIELDS}
    lines = [
        f"Период: {len(full_days)} полных суток",
        f"Конв.4={fmt(s['kv4'])}т Конв.3={fmt(s['kv3'])}т",
        f"Баланс1={sign(bal1(s))} БалансС1={sign(balc1(s))} "
        f"Баланс2={sign(bal2(s))} БалансС2={sign(balc2(s))}",
        "",
        "Нарушения по дням:",
    ]
    found = False
    for r in full_days:
        al = build_alerts(r, label=f"д.{r['day_num']}")
        if al:
            found = True
            lines.extend(a[1] for a in al)
    if not found:
        lines.append("Нарушений не обнаружено.")
    return "\n".join(lines)


# ════════════════════════════════════════════════════════
#  TELEGRAM BOT
# ════════════════════════════════════════════════════════
bot = Bot(token=BOT_TOKEN)
dp = Dispatcher(storage=MemoryStorage())


class AIState(StatesGroup):
    waiting_for_question = State()


class StockInput(StatesGroup):
    waiting_for_prev = State()   # запас за день N-1
    waiting_for_curr = State()   # запас за день N


class StockNightInput(StatesGroup):
    waiting_for_night = State()   # запас на конец ночи (до начала смены 1)
    waiting_for_morning = State() # запас на утро (после смены 1)


def main_keyboard() -> ReplyKeyboardMarkup:
    return ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text="📅 Суточный баланс"), KeyboardButton(text="🌙 Ночная смена")],
            [KeyboardButton(text="📆 Недельная сводка"), KeyboardButton(text="🗓 Месячный итог")],
            [KeyboardButton(text="🔔 Алерты"), KeyboardButton(text="🔍 Дублирование весов")],
            [KeyboardButton(text="🏭 Склад концентрата"), KeyboardButton(text="📥 Ввести запас")],
            [KeyboardButton(text="🌅 Склад — ночная смена"), KeyboardButton(text="❓ Помощь")],
            [KeyboardButton(text="🤖 Спросить AI")],
        ],
        resize_keyboard=True,
    )


@dp.message(Command("start"))
async def cmd_start(msg: Message):
    init_db()
    await msg.answer(
        "⚖️ *ДОФ Баланс*\n"
        "Мониторинг конвейерных весов · 2 очереди производства\n\n"
        "📁 Просто отправьте файл отчёта *report.xls* — бот автоматически "
        "прочитает данные, посчитает балансы и проверит технологические нормы.",
        parse_mode="Markdown",
        reply_markup=main_keyboard(),
    )


@dp.message(F.document)
async def handle_report(msg: Message):
    doc: Document = msg.document
    fn = doc.file_name or ""
    if not (fn.lower().endswith(".xls") or fn.lower().endswith(".xlsx")):
        await msg.answer("❌ Поддерживаются только файлы .xls или .xlsx")
        return

    wait_msg = await msg.answer("⏳ Читаю файл...")

    try:
        file_obj = await bot.get_file(doc.file_id)
        buf = io.BytesIO()
        await bot.download_file(file_obj.file_path, buf)
        file_bytes = buf.getvalue()
    except Exception as e:
        await wait_msg.edit_text(f"❌ Не удалось скачать файл: {e}")
        return

    parsed = parse_report(file_bytes, fn)

    if parsed.get("error") or not parsed["daily_by_shift"]:
        err = parsed.get("error", "Неизвестная ошибка структуры файла.")
        await wait_msg.edit_text(f"⚠️ {err}")
        return

    now = datetime.now()
    saved, max_day = db_save_daily(parsed, msg.from_user.id, fn, now.year, now.month)

    if saved == 0:
        await wait_msg.edit_text(
            "⚠️ Файл прочитан, но в нём нет ни одних полных суток "
            "(нужны минимум 2 дня с обеими сменами)."
        )
        return

    await wait_msg.edit_text(
        f"✅ *Отчёт загружен!*\n\n"
        f"📊 Лист: `{parsed['used_sheet']}`\n"
        f"📅 Период: {parsed.get('period') or 'текущий месяц'}\n"
        f"💾 В баланс включено полных суток: {saved}\n"
        f"🌙 День {max_day}: исключён из баланса целиком "
        f"(Смена 2 ещё идёт) — Смена 1 доступна отдельно по кнопке\n\n"
        f"Используйте кнопки меню для отчётов.",
        parse_mode="Markdown",
    )


# ── СУТОЧНЫЙ БАЛАНС (полные сутки) ──────────────────────
@dp.message(F.text == "📅 Суточный баланс")
async def report_daily(msg: Message):
    now = datetime.now()
    rows = db_get_month_data(now.year, now.month)
    full = rows  # все записи в daily_data уже полные сутки
    if not full:
        await msg.answer("📭 Нет завершённых суток. Загрузите файл отчёта.")
        return

    lines = [f"📊 *Суточный баланс ({now.month:02d}/{now.year})*", "_(только завершённые сутки)_\n"]
    lines.append("`Дн  Б1      Б2      Кв4     Кв3`")
    for r in full:
        b1, b2 = bal1(r), bal2(r)
        lines.append(
            f"`{r['day_num']:>2d}  "
            f"{em_bal(b1)}{sign(b1,1):>6s}  "
            f"{em_bal(b2)}{sign(b2,1):>6s}  "
            f"{fmt2(r.get('kv4')):>6s}  "
            f"{fmt2(r.get('kv3')):>6s}`"
        )
    await msg.answer("\n".join(lines)[:4000], parse_mode="Markdown")


# ── НОЧНАЯ СМЕНА (Смена 1 незавершённого дня) ────────────
@dp.message(F.text == "🌙 Ночная смена")
async def report_night_shift(msg: Message):
    now = datetime.now()
    ns = db_get_night_shift(now.year, now.month)
    if not ns:
        await msg.answer(
            "📋 Нет данных по ночной смене.\n"
            "Загрузите свежий report.xls."
        )
        return

    day_num = ns["day_num"]
    b1, b2 = bal1(ns), bal2(ns)
    bc1, bc2 = balc1(ns), balc2(ns)

    text = (
        f"🌙 *Ночная смена — день {day_num:02d}.{now.month:02d}*\n"
        f"_Смена 1 (19:30–07:30), отдельно от месячного баланса_\n\n"
        f"━━ 1 очередь (Кв4={fmt(ns.get('kv4'))}т) ━━\n"
        f"{em_bal(b1)} Баланс 1: {sign(b1)}\n"
        f"{em_bal(bc1)} Баланс С.1: {sign(bc1)}\n"
        f"Конв.14: {fmt(ns.get('kv14'))}т  Конв.102: {fmt(ns.get('kv102'))}т\n\n"
        f"━━ 2 очередь (Кв3={fmt(ns.get('kv3'))}т) ━━\n"
        f"{em_bal(b2)} Баланс 2: {sign(b2)}\n"
        f"{em_bal(bc2)} Баланс С.2: {sign(bc2)}\n"
        f"Конв.15: {fmt(ns.get('kv15'))}т  Конв.101: {fmt(ns.get('kv101'))}т\n\n"
        f"_Смена 2 этого дня ещё идёт — будет учтена завтра, "
        f"когда сутки станут полными._"
    )
    await msg.answer(text, parse_mode="Markdown")


# ── НЕДЕЛЬНАЯ СВОДКА ─────────────────────────────────────
@dp.message(F.text == "📆 Недельная сводка")
async def report_weekly(msg: Message):
    now = datetime.now()
    rows = db_get_month_data(now.year, now.month)
    full = rows  # все записи в daily_data уже полные сутки
    if not full:
        await msg.answer("📭 Нет завершённых суток.")
        return

    last7 = full[-7:]
    s = {k: sum(r.get(k, 0) for r in last7) for k in FIELDS}
    lines = [f"📆 *Сводка за {len(last7)} последних суток*\n"]
    lines.append(f"{em_bal(bal1(s))} Баланс 1: {sign(bal1(s))}")
    lines.append(f"{em_bal(bal2(s))} Баланс 2: {sign(bal2(s))}")
    lines.append(f"Конв.4: {fmt(s['kv4'])}т   Конв.3: {fmt(s['kv3'])}т\n")
    for r in last7:
        lines.append(f"▪️ День {r['day_num']}: Кв4={fmt(r.get('kv4'))}т  Кв3={fmt(r.get('kv3'))}т")
    await msg.answer("\n".join(lines)[:4000], parse_mode="Markdown")


# ── МЕСЯЧНЫЙ ИТОГ ─────────────────────────────────────────
@dp.message(F.text == "🗓 Месячный итог")
async def report_monthly(msg: Message):
    now = datetime.now()
    rows = db_get_month_data(now.year, now.month)
    full = rows  # все записи в daily_data уже полные сутки
    if not full:
        await msg.answer("📭 Нет полных данных за месяц.")
        return

    s = {k: sum(r.get(k, 0) for r in full) for k in FIELDS}
    b1, b2 = bal1(s), bal2(s)
    bc1, bc2 = balc1(s), balc2(s)

    text = (
        f"🗓 *Месячный итог ({now.month:02d}/{now.year})*\n"
        f"_Учтено суток: {len(full)}, текущая смена исключена_\n\n"
        f"━━ 1 очередь ━━\n"
        f"Конв.4: {fmt(s['kv4'])}т (осн.)   Конв.4Д: {fmt(s['kv4d'])}т (дубл.)\n"
        f"Конв.14: {fmt(s['kv14'])}т  ({pct(s['kv14'],s['kv4']):.0f}%)\n"
        f"Конв.32: {fmt(s['kv32'])}т  ({pct(s['kv32'],s['kv4']):.0f}%)\n"
        f"Конв.34: {fmt(s['kv34'])}т  ({pct(s['kv34'],s['kv4']):.0f}%)\n"
        f"Конв.102: {fmt(s['kv102'])}т  ({pct(s['kv102'],s['kv4']):.0f}%)\n"
        f"{em_bal(b1)} Баланс 1: {sign(b1)}\n"
        f"{em_bal(bc1)} Баланс С.1: {sign(bc1)}\n\n"
        f"━━ 2 очередь ━━\n"
        f"Конв.3: {fmt(s['kv3'])}т (осн.)   Конв.3Д: {fmt(s['kv3d'])}т (дубл.)\n"
        f"Конв.15: {fmt(s['kv15'])}т  ({pct(s['kv15'],s['kv3']):.0f}%)\n"
        f"Конв.33: {fmt(s['kv33'])}т  ({pct(s['kv33'],s['kv3']):.0f}%)\n"
        f"Конв.101: {fmt(s['kv101'])}т  ({pct(s['kv101'],s['kv3']):.0f}%)\n"
        f"{em_bal(b2)} Баланс 2: {sign(b2)}\n"
        f"{em_bal(bc2)} Баланс С.2: {sign(bc2)}"
    )
    await msg.answer(text, parse_mode="Markdown")


# ── АЛЕРТЫ ────────────────────────────────────────────────
@dp.message(F.text == "🔔 Алерты")
async def report_alerts(msg: Message):
    now = datetime.now()
    rows = db_get_month_data(now.year, now.month)
    full = rows  # все записи в daily_data уже полные сутки
    if not full:
        await msg.answer("📭 Нет завершённых суток.")
        return

    all_alerts = []
    for r in full:
        all_alerts.extend(build_alerts(r, label=f"д.{r['day_num']}"))

    # Алерты по складу концентрата
    stock_history = db_get_stock_history(now.year, now.month)
    stock_alerts = []
    for r in stock_history:
        stock_curr = r.get("stock_curr", 0)
        nesovp     = r.get("nesovpadenie", 0)
        day        = r["day_num"]
        if stock_curr < 0:
            stock_alerts.append(("crit", f"🚨 Склад д.{day}: запас отрицательный ({fmt(stock_curr)}т)!"))
        elif abs(nesovp) > STOCK_DIFF_WARN:
            em = "crit" if abs(nesovp) > STOCK_DIFF_WARN * 2 else "warn"
            stock_alerts.append((em, f"{'🚨' if em=='crit' else '⚠️'} Склад д.{day}: несовпадение {fmt(nesovp)}т"))

    has_any = all_alerts or stock_alerts
    if not has_any:
        await msg.answer("✅ *Нарушений не обнаружено за весь период.*", parse_mode="Markdown")
        return

    crits = [a[1] for a in all_alerts if a[0] == "crit"]
    warns = [a[1] for a in all_alerts if a[0] == "warn"]
    s_crits = [a[1] for a in stock_alerts if a[0] == "crit"]
    s_warns = [a[1] for a in stock_alerts if a[0] == "warn"]

    total_c = len(crits) + len(s_crits)
    total_w = len(warns) + len(s_warns)
    lines = [f"🔔 *Алерты: {total_c} крит. / {total_w} предупр.*\n"]
    if crits or s_crits:
        lines.append("🚨 *Критичные:*")
        lines.extend(crits + s_crits)
    if warns or s_warns:
        lines.append("\n⚠️ *Предупреждения:*")
        lines.extend(warns + s_warns)
    await msg.answer("\n".join(lines)[:4000], parse_mode="Markdown")


# ── ДУБЛИРОВАНИЕ ВЕСОВ ───────────────────────────────────
@dp.message(F.text == "🔍 Дублирование весов")
async def report_doubles(msg: Message):
    now = datetime.now()
    rows = db_get_month_data(now.year, now.month)
    full = rows  # все записи в daily_data уже полные сутки
    if not full:
        await msg.answer("📭 Нет завершённых суток.")
        return

    lines = ["🔍 *Расхождение основных и дублирующих весов*", "_норма <0.5%, критично >2%_\n"]
    lines.append("`Дн  Кв4 vs 4Д     Кв3 vs 3Д`")
    for r in full:
        st4, p4, _ = check_doubles(r.get("kv4", 0), r.get("kv4d", 0))
        st3, p3, _ = check_doubles(r.get("kv3", 0), r.get("kv3d", 0))
        lines.append(f"`{r['day_num']:>2d}   {em_dup(st4)}{p4:>5.2f}%      {em_dup(st3)}{p3:>5.2f}%`")
    await msg.answer("\n".join(lines)[:4000], parse_mode="Markdown")


# ── СКЛАД ВЛАЖНОГО КОНЦЕНТРАТА ────────────────────────────
@dp.message(F.text == "🏭 Склад концентрата")
async def report_stock(msg: Message):
    now = datetime.now()
    history = db_get_stock_history(now.year, now.month)
    if not history:
        await msg.answer(
            "📭 Нет данных по складу за текущий месяц.\n"
            "Используйте *📥 Ввести запас* чтобы добавить данные.",
            parse_mode="Markdown"
        )
        return

    lines = [
        f"🏭 *Склад влажного концентрата · {now.month:02d}/{now.year}*\n",
        "`Дн  Произв.  Отгруж.  Вес.изм  Мркш.изм  Несовп`",
    ]
    for r in history:
        nesovp = r.get("nesovpadenie", 0)
        em = "🚨" if abs(nesovp) > STOCK_DIFF_WARN else "✅"
        lines.append(
            f"`{r['day_num']:>2d}  "
            f"{fmt2(r.get('produced',0)):>7s}  "
            f"{fmt2(r.get('shipped',0)):>7s}  "
            f"{fmt2(r.get('ves_izm',0)):>+7s}  "
            f"{fmt2(r.get('marksh_izm',0)):>+7s}  "
            f"{em}{nesovp:+.0f}`"
        )

    last = history[-1]
    lines.append(
        f"\n📅 Последний расчёт: день {last['day_num']}\n"
        f"⚙️ Произведено:  {fmt(last.get('produced',0))} т\n"
        f"🚂 Отгружено:    {fmt(last.get('shipped',0))} т\n"
        f"⚖️ Вес.изм.:     {fmt(last.get('ves_izm',0))} т\n"
        f"📐 Маркш.изм.:   {fmt(last.get('marksh_izm',0))} т\n"
        f"❗ Несовпадение: {fmt(last.get('nesovpadenie',0))} т"
    )
    await msg.answer("\n".join(lines)[:4000], parse_mode="Markdown")


# ── ВВОД ЗАПАСОВ СКЛАДА (2 числа: за N-1 и за N) ─────────
@dp.message(F.text == "📥 Ввести запас")
async def stock_input_start(msg: Message, state: FSMContext):
    now = datetime.now()
    rows = db_get_month_data(now.year, now.month)
    if not rows:
        await msg.answer(
            "📭 Нет данных из отчёта. Сначала загрузите *report.xls*.",
            parse_mode="Markdown"
        )
        return

    # Последний полный день в отчёте (напр. день 18 при скачивании 19-го)
    day_num = rows[-1]["day_num"]
    await state.update_data(day_num=day_num, year=now.year, month=now.month)
    await state.set_state(StockInput.waiting_for_prev)

    await msg.answer(
        f"📥 *Склад влажного концентрата — день {day_num:02d}.{now.month:02d}*\n\n"
        f"Шаг 1 из 2\n"
        f"Введите запас на складе на конец *{day_num-1:02d}.{now.month:02d}* (тонн):",
        parse_mode="Markdown"
    )


@dp.message(StockInput.waiting_for_prev)
async def stock_input_prev(msg: Message, state: FSMContext):
    try:
        stock_prev = float(msg.text.strip().replace(",", ".").replace(" ", ""))
    except ValueError:
        await msg.answer("❌ Введите число (тонн). Например: `61500`", parse_mode="Markdown")
        return

    data = await state.get_data()
    await state.update_data(stock_prev=stock_prev)
    await state.set_state(StockInput.waiting_for_curr)

    day_num = data["day_num"]
    now_month = data["month"]
    await msg.answer(
        f"Шаг 2 из 2\n"
        f"Введите запас на складе на конец *{day_num:02d}.{now_month:02d}* (тонн):",
        parse_mode="Markdown"
    )


@dp.message(StockInput.waiting_for_curr)
async def stock_input_curr(msg: Message, state: FSMContext):
    try:
        stock_curr = float(msg.text.strip().replace(",", ".").replace(" ", ""))
    except ValueError:
        await msg.answer("❌ Введите число (тонн). Например: `52500`", parse_mode="Markdown")
        return

    data = await state.get_data()
    await state.clear()
    year     = data["year"]
    month    = data["month"]
    day_num  = data["day_num"]
    stock_prev = data["stock_prev"]

    result = db_save_stock(year, month, day_num, stock_prev, stock_curr, msg.from_user.id)

    ves_izm    = result["ves_izm"]
    marksh_izm = result["marksh_izm"]
    nesovp     = result["nesovpadenie"]

    lines = [
        f"✅ *Расчёт склада — день {day_num:02d}.{month:02d}*\n",
        f"📦 Запас на {day_num-1:02d}.{month:02d}: {fmt(stock_prev)} т",
        f"📦 Запас на {day_num:02d}.{month:02d}: {fmt(stock_curr)} т",
        f"",
        f"⚙️ Произведено:   {fmt(result['produced'])} т",
        f"🚂 Отгружено:     {fmt(result['shipped'])} т",
        f"⚖️ Вес.изменение: {fmt(ves_izm)} т",
        f"📐 Маркш.изм.:    {fmt(marksh_izm)} т",
        f"❗ Несовпадение:  {fmt(nesovp)} т",
    ]

    if stock_curr < 0:
        lines.append("\n🚨 *АЛЕРТ: запас отрицательный! Проверьте данные.*")
    elif abs(nesovp) > STOCK_DIFF_WARN:
        em = "🚨" if abs(nesovp) > STOCK_DIFF_WARN * 2 else "⚠️"
        lines.append(
            f"\n{em} *Несовпадение {fmt(abs(nesovp))} т > 500 т!*\n"
            f"Проверьте весы 44/46/74 и 65/66/84."
        )
    else:
        lines.append("\n✅ Несовпадение в норме (< 500 т)")

    await msg.answer("\n".join(lines), parse_mode="Markdown")


# ── СКЛАД — НОЧНАЯ СМЕНА (утренний расчёт по Смене 1) ────
@dp.message(F.text == "🌅 Склад — ночная смена")
async def stock_night_start(msg: Message, state: FSMContext):
    now = datetime.now()
    ns = db_get_night_shift(now.year, now.month)
    if not ns:
        await msg.answer(
            "📭 Нет данных по ночной смене.\n"
            "Сначала загрузите *report.xls*.",
            parse_mode="Markdown"
        )
        return

    day_num  = ns["day_num"]
    produced = calc_produced(ns)
    shipped  = calc_shipped(ns)

    await state.update_data(
        day_num=day_num,
        year=now.year,
        month=now.month,
        produced=produced,
        shipped=shipped,
    )
    await state.set_state(StockNightInput.waiting_for_night)

    await msg.answer(
        f"🌅 *Склад — ночная смена {day_num:02d}.{now.month:02d}*\n\n"
        f"_Смена 1 (19:30–07:30) из отчёта:_\n"
        f"⚙️ Произведено: {fmt(produced)} т\n"
        f"🚂 Отгружено:   {fmt(shipped)} т\n\n"
        f"Шаг 1 из 2\n"
        f"Введите запас на складе на *начало смены* (вечер {day_num-1:02d}.{now.month:02d}, тонн):",
        parse_mode="Markdown"
    )


@dp.message(StockNightInput.waiting_for_night)
async def stock_night_prev(msg: Message, state: FSMContext):
    try:
        stock_night = float(msg.text.strip().replace(",", ".").replace(" ", ""))
    except ValueError:
        await msg.answer("❌ Введите число (тонн). Например: `61500`", parse_mode="Markdown")
        return

    data = await state.get_data()
    await state.update_data(stock_night=stock_night)
    await state.set_state(StockNightInput.waiting_for_morning)

    day_num = data["day_num"]
    month   = data["month"]
    await msg.answer(
        f"Шаг 2 из 2\n"
        f"Введите запас на складе на *конец смены* (утро {day_num:02d}.{month:02d}, тонн):",
        parse_mode="Markdown"
    )


@dp.message(StockNightInput.waiting_for_morning)
async def stock_night_curr(msg: Message, state: FSMContext):
    try:
        stock_morning = float(msg.text.strip().replace(",", ".").replace(" ", ""))
    except ValueError:
        await msg.answer("❌ Введите число (тонн). Например: `58000`", parse_mode="Markdown")
        return

    data = await state.get_data()
    await state.clear()

    day_num     = data["day_num"]
    month       = data["month"]
    produced    = data["produced"]
    shipped     = data["shipped"]
    stock_night = data["stock_night"]

    ves_izm    = produced - shipped
    marksh_izm = stock_morning - stock_night
    nesovp     = ves_izm - marksh_izm

    lines = [
        f"🌅 *Склад — ночная смена {day_num:02d}.{month:02d}*\n",
        f"📦 Запас на начало смены: {fmt(stock_night)} т",
        f"📦 Запас на конец смены:  {fmt(stock_morning)} т",
        f"",
        f"⚙️ Произведено:   {fmt(produced)} т",
        f"🚂 Отгружено:     {fmt(shipped)} т",
        f"⚖️ Вес.изменение: {fmt(ves_izm)} т",
        f"📐 Маркш.изм.:    {fmt(marksh_izm)} т",
        f"❗ Несовпадение:  {fmt(nesovp)} т",
    ]

    if stock_morning < 0:
        lines.append("\n🚨 *АЛЕРТ: запас отрицательный! Проверьте данные.*")
    elif abs(nesovp) > STOCK_DIFF_WARN:
        em = "🚨" if abs(nesovp) > STOCK_DIFF_WARN * 2 else "⚠️"
        lines.append(
            f"\n{em} *Несовпадение {fmt(abs(nesovp))} т > 500 т!*\n"
            f"Проверьте весы 44/46/74 и 65/66/84."
        )
    else:
        lines.append("\n✅ Несовпадение в норме (< 500 т)")

    await msg.answer("\n".join(lines), parse_mode="Markdown")


# ── AI ────────────────────────────────────────────────────
@dp.message(F.text == "🤖 Спросить AI")
async def ai_request(msg: Message, state: FSMContext):
    await state.set_state(AIState.waiting_for_question)
    await msg.answer(
        "🤖 *AI-Агент метролога ДОФ*\nВведите вопрос по балансу, нормам или диагностике весов:",
        parse_mode="Markdown",
    )


@dp.message(AIState.waiting_for_question)
async def ai_processing(msg: Message, state: FSMContext):
    await state.clear()
    wait = await msg.answer("⏳ Анализирую...")
    now = datetime.now()
    rows = db_get_month_data(now.year, now.month)
    context = make_ai_context(rows) if rows else "Нет данных."
    answer = await ask_ai(msg.text, context)
    await wait.edit_text(f"🤖 *AI-Агент:*\n\n{answer[:4000]}", parse_mode="Markdown")


# ── ПОМОЩЬ ────────────────────────────────────────────────
@dp.message(F.text == "❓ Помощь")
async def h_help(msg: Message):
    await msg.answer(
        "📖 *Как пользоваться*\n\n"
        "1️⃣ Отправьте файл `report.xls` боту\n"
        "2️⃣ Бот прочитает все данные автоматически\n"
        "3️⃣ Последний день файла ПОЛНОСТЬЮ исключается из баланса "
        "(Смена 2 ещё идёт) — учитываются только завершённые сутки\n\n"
        "*Кнопки:*\n"
        "📅 Суточный баланс — таблица по дням (только полные сутки)\n"
        "🌙 Ночная смена — Смена 1 текущего незавершённого дня, отдельным балансом\n"
        "📆 Недельная сводка — за последние 7 полных суток\n"
        "🗓 Месячный итог — накопительный баланс (без текущего дня)\n"
        "🔔 Алерты — все нарушения норм и балансов\n"
        "🔍 Дублирование весов — Конв.4/4Д и Конв.3/3Д\n"
        "🤖 Спросить AI — диагностика и рекомендации\n\n"
        "*Нормы конвейеров:*\n"
        "Конв.14/15: 60–80%  |  Конв.101/102: 6–20%\n"
        "Конв.19/31/32: 40–60%  |  Конв.33/34: 83–87%\n"
        "Конв.10/34А: 15–45%\n\n"
        "*Баланс:* ✅ ±2%  ⚠️ ±5%  🚨 >5%\n"
        "*Дублирование:* ✅ <0.5%  ⚠️ <2%  🚨 >2%",
        parse_mode="Markdown",
    )


# ════════════════════════════════════════════════════════
#  ЗАПУСК
# ════════════════════════════════════════════════════════
async def main():
    init_db()
    logger.info("ДОФ Баланс Bot запущен.")
    await dp.start_polling(bot)


if __name__ == "__main__":
    asyncio.run(main())