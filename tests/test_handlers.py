import io
import os
import tempfile
import unittest
from datetime import date
from pathlib import Path
from types import SimpleNamespace
from unittest.mock import AsyncMock, patch

import openpyxl

os.environ.setdefault(
    "BOT_TOKEN",
    "123456789:ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmno",
)

import main


class FakeSentMessage:
    def __init__(self, text):
        self.text = text
        self.edits = []

    async def edit_text(self, text, **kwargs):
        self.edits.append((text, kwargs))


class FakeMessage:
    def __init__(self, *, user_id=101, text=None, document=None, from_user=True):
        self.from_user = SimpleNamespace(id=user_id) if from_user else None
        self.text = text
        self.document = document
        self.answers = []
        self.sent_messages = []

    async def answer(self, text, **kwargs):
        self.answers.append((text, kwargs))
        sent = FakeSentMessage(text)
        self.sent_messages.append(sent)
        return sent


class FakeState:
    def __init__(self):
        self.data = {}
        self.state = None
        self.clear_count = 0

    async def clear(self):
        self.data = {}
        self.state = None
        self.clear_count += 1

    async def update_data(self, **kwargs):
        self.data.update(kwargs)

    async def get_data(self):
        return dict(self.data)

    async def set_state(self, state):
        self.state = state


class FakeAIResponse:
    def __init__(self, status=200, payload=None, body=""):
        self.status = status
        self.payload = payload
        self.body = body

    async def __aenter__(self):
        return self

    async def __aexit__(self, _exc_type, _exc, _traceback):
        return False

    async def json(self):
        return self.payload

    async def text(self):
        return self.body


class FakeAISession:
    def __init__(self, response):
        self.response = response
        self.request = None

    async def __aenter__(self):
        return self

    async def __aexit__(self, _exc_type, _exc, _traceback):
        return False

    def post(self, url, **kwargs):
        self.request = (url, kwargs)
        return self.response


def production_values():
    values = {field: 0.0 for field in main.FIELDS}
    values.update(
        {
            "kv4": 1000,
            "kv4d": 1000,
            "kv14": 700,
            "kv32": 550,
            "kv34": 850,
            "kv34a": 300,
            "kv102": 100,
            "kv24p": 50,
            "kv24hv": 50,
            "kv28a1": 50,
            "kv3": 1000,
            "kv3d": 1000,
            "kv15": 700,
            "kv19": 500,
            "kv31": 550,
            "kv33": 850,
            "kv101": 150,
            "kv28a2": 0,
            "kv44": 100,
            "kv44d": 100,
            "kv46d": 50,
            "kv74": 100,
            "kv74d": 100,
            "kv65mps": 10,
            "kv65cpo": 10,
            "kv66mps": 10,
            "kv66cpo": 10,
            "kv84mps": 10,
            "kv84cpo": 10,
        }
    )
    return values


def two_shifts(values):
    return {
        1: {field: amount / 2 for field, amount in values.items()},
        2: {field: amount / 2 for field, amount in values.items()},
    }


def upload_workbook_bytes():
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "ДетСменТекМесяц"
    sheet.cell(1, 1, "Отчёт за июль 2026")
    sheet.cell(3, 1, "Дата:")
    sheet.cell(4, 1, "Смена:")
    sheet.cell(3, 2, date(2026, 7, 1))
    sheet.cell(4, 2, 1)
    sheet.cell(4, 3, 2)
    sheet.cell(5, 1, "Конвейер 4")
    sheet.cell(5, 2, 500)
    sheet.cell(5, 3, 500)
    sheet.cell(6, 1, "Конвейер 3")
    sheet.cell(6, 2, 500)
    sheet.cell(6, 3, 500)
    buffer = io.BytesIO()
    workbook.save(buffer)
    workbook.close()
    return buffer.getvalue()


class HandlerSmokeTests(unittest.IsolatedAsyncioTestCase):
    async def asyncSetUp(self):
        self.temp_dir = tempfile.TemporaryDirectory(prefix="dof-handler-test-")
        self.original_db_path = main.DB_PATH
        main.DB_PATH = str(Path(self.temp_dir.name) / "handlers.db")
        main.init_db()

        values = production_values()
        main.db_save_daily(
            {
                "period": "Отчёт за июль 2026",
                "daily_by_shift": {
                    1: two_shifts(values),
                    2: {
                        1: {field: amount / 2 for field, amount in values.items()},
                        2: {},
                    },
                },
            },
            101,
            "active.xlsx",
            2026,
            7,
            today=date(2026, 7, 2),
        )
        main.db_save_stock(2026, 7, 1, 1000, 1100, 101)

    async def asyncTearDown(self):
        main.DB_PATH = self.original_db_path
        self.temp_dir.cleanup()

    async def test_all_read_only_menu_handlers_render(self):
        handlers_and_phrases = (
            (main.report_daily, "Суточный баланс"),
            (main.report_night_shift, "Ночная смена"),
            (main.report_sliding_balance, "Скользящий баланс"),
            (main.report_weekly, "Сводка за"),
            (main.report_monthly, "Месячный итог"),
            (main.report_alerts, "Трёхсуточный сигнал"),
            (main.report_doubles, "Расхождение основных"),
            (main.report_stock, "Склад влажного концентрата"),
            (main.h_help, "Как пользоваться"),
        )

        for handler, expected in handlers_and_phrases:
            with self.subTest(handler=handler.__name__):
                message = FakeMessage(user_id=101)
                state = FakeState()
                await handler(message, state)
                rendered = "\n".join(text for text, _kwargs in message.answers)
                self.assertIn(expected, rendered)
                self.assertGreaterEqual(state.clear_count, 1)

    async def test_stock_input_dialog_saves_finite_numbers(self):
        state = FakeState()
        await main.stock_input_start(FakeMessage(user_id=101), state)
        self.assertEqual(state.state, main.StockInput.waiting_for_prev)

        await main.stock_input_prev(
            FakeMessage(user_id=101, text="1 200,5"),
            state,
        )
        self.assertEqual(state.state, main.StockInput.waiting_for_curr)

        finish = FakeMessage(user_id=101, text="1300,5")
        await main.stock_input_curr(finish, state)
        self.assertIsNone(state.state)
        self.assertIn("Расчёт склада", finish.answers[0][0])

    async def test_night_stock_dialog_renders_result(self):
        state = FakeState()
        await main.stock_night_start(FakeMessage(user_id=101), state)
        self.assertEqual(state.state, main.StockNightInput.waiting_for_night)

        await main.stock_night_prev(FakeMessage(user_id=101, text="1000"), state)
        finish = FakeMessage(user_id=101, text="1050")
        await main.stock_night_curr(finish, state)

        self.assertIn("Склад — ночная смена", finish.answers[0][0])
        self.assertIsNone(state.state)

    async def test_ai_dialog_splits_long_response_without_markdown_parsing(self):
        state = FakeState()
        start = FakeMessage(user_id=101)
        await main.ai_request(start, state)
        self.assertEqual(state.state, main.AIState.waiting_for_question)

        question = FakeMessage(user_id=101, text="Проверь баланс")
        with patch.object(main, "ask_ai", AsyncMock(return_value="А" * 8000)):
            await main.ai_processing(question, state)

        self.assertGreaterEqual(len(question.answers), 2)
        self.assertTrue(question.sent_messages[0].edits)
        self.assertTrue(
            all(
                kwargs.get("parse_mode") is None
                for _text, kwargs in question.answers[1:]
            )
        )

    async def test_access_middleware_blocks_channel_and_unlisted_user(self):
        middleware = main.AccessMiddleware()
        handler = AsyncMock(return_value="handled")

        channel_message = FakeMessage(from_user=False)
        self.assertIsNone(await middleware(handler, channel_message, {}))
        handler.assert_not_awaited()

        denied_message = FakeMessage(user_id=999)
        with patch.object(main, "ALLOWED_USER_IDS", {101}):
            self.assertIsNone(await middleware(handler, denied_message, {}))
        handler.assert_not_awaited()

        allowed_message = FakeMessage(user_id=101)
        with patch.object(main, "ALLOWED_USER_IDS", {101}):
            self.assertEqual(
                await middleware(handler, allowed_message, {}),
                "handled",
            )
        handler.assert_awaited_once()

    async def test_successful_document_upload_updates_active_report(self):
        file_bytes = upload_workbook_bytes()
        document = SimpleNamespace(
            file_name="report_july_2026.xlsx",
            file_size=len(file_bytes),
            file_id="telegram-file-id",
        )
        message = FakeMessage(user_id=101, document=document)
        state = FakeState()

        async def download_file(_path, buffer):
            buffer.write(file_bytes)

        fake_bot = SimpleNamespace(
            get_file=AsyncMock(return_value=SimpleNamespace(file_path="file.xlsx")),
            download_file=download_file,
        )
        with patch.object(main, "bot", fake_bot):
            await main.handle_report(message, state)

        self.assertEqual(main.db_get_active_report(101)["filename"], document.file_name)
        self.assertIn("Отчёт загружен", message.sent_messages[0].edits[-1][0])


class AIClientTests(unittest.IsolatedAsyncioTestCase):
    async def test_ai_is_explicitly_disabled_without_api_key(self):
        with patch.object(main, "ANTHROPIC_API_KEY", None):
            answer = await main.ask_ai("вопрос", "контекст")
        self.assertIn("не задан ANTHROPIC_API_KEY", answer)

    async def test_ai_text_blocks_and_length_stop_are_handled(self):
        response = FakeAIResponse(
            payload={
                "content": [
                    {"type": "text", "text": "Первая часть"},
                    {"type": "tool_use", "name": "ignored"},
                    {"type": "text", "text": "Вторая часть"},
                ],
                "stop_reason": "max_tokens",
            }
        )
        session = FakeAISession(response)
        with (
            patch.object(main, "ANTHROPIC_API_KEY", "test-key"),
            patch.object(main.aiohttp, "ClientSession", return_value=session),
        ):
            answer = await main.ask_ai("вопрос", "контекст")

        self.assertIn("Первая часть\nВторая часть", answer)
        self.assertIn("лимиту длины", answer)
        self.assertEqual(session.request[0], "https://api.anthropic.com/v1/messages")
        self.assertEqual(session.request[1]["json"]["model"], main.AI_MODEL)

    async def test_ai_unexpected_json_does_not_crash_handler(self):
        session = FakeAISession(FakeAIResponse(payload=["unexpected"]))
        with (
            patch.object(main, "ANTHROPIC_API_KEY", "test-key"),
            patch.object(main.aiohttp, "ClientSession", return_value=session),
        ):
            answer = await main.ask_ai("вопрос", "контекст")

        self.assertIn("неизвестного формата", answer)


if __name__ == "__main__":
    unittest.main()
