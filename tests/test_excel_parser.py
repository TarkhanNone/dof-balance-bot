import base64
import io
import os
import unittest
import zlib
from datetime import date
from unittest.mock import patch

import openpyxl

os.environ.setdefault(
    "BOT_TOKEN",
    "123456789:ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmno",
)

import main

# Небольшой настоящий BIFF8-файл, созданный Excel-совместимым генератором.
# Он хранится сжатым в исходнике, чтобы тест пути .xls не требовал xlwt.
XLS_FIXTURE = zlib.decompress(
    base64.b64decode(
        "eNrtWE9IVEEc/ubt7K67+Ge1NdBi2YSszMOmZWbpqpBeyrRLEUFpeghljaVLHsoyj0HQqegieOlidekPFdStQ2DYIQiCtY6dgoIOuq9vfvuef2oPLphUvO8x38z85jffmzczb2beeztbnpl6WD2PX9AGH7J2CIEVNsUQcjMRsNy2TdKNixhsD/8UQkUcyIAfz0reBM0YmvGeh4UH+hUZ+MRwGhfQM5oaim8gOqUN/cq0oZWscJeWUlRJqyqEzwlvEr4vns+F28VyQ7iVvhl1CrPJnrpmZxaftGqkrBRG97HU+SCWPajEazOLr9xUOV8/OtLn+0f+zoKYLsY0OG7dQ6mhdP9IBlEO4DS+23Hgm/umvox79o21K9D+Y7U9mMd+y9LAOOyzMsEnUYsmbUrKVFTv0516mz7EuFXXkFv0VnKHPqa79AIucQpAQgVUjK7duo8cx36dIDfrHn1QH2GqAQmGJrpSM0GXhG5hG5UrbbJhqGrdxkwDTQcY2llxb35zI83cBWTViKxaNUrkbSomD6JM0uXyTkW4Lyzc+zp3dKA3eUYs47JT5PaT7aYLYOOqqcHKpXD1co9XJzV2C18T1S2SrhaO0odxbW+lk+iaEJ/rUlrL+zQK3iV3rEjvZHryS9+T2OTn5C6mZ7rnx6Iz75NTqOH+Nsj65ppAvapXd24bPE26sXLWno/CVb+tQ0VWxHk229k0y7AofQb2BxwPS55utYclHqZPLlthp//GTLEVYM5iLmf3mdSSXTNnFH15FH2iaHRfIMacUQyILSy8nNaiofNoaPHxORpaNOaCWOKBkMs5DX8eDb9oaEfDLxqdASxxOOiyRptVgUemyVzFlxGGBw8ePHjw4MGDh/WBOW+a45ZPToG5M7c5Fwad/zqLDFnvN8l/i+MY5XWRH6aHkWKc5qdlIdgMv3K11BrruP8LDU7w7mkMY0DaMVzw/OXXmFr5PGuuGFm/V6jQ+2cLaecfvv9PwLPV8A=="
    )
)


def make_xlsx(
    *,
    sheet_name=" ДетСменТекМесяц ",
    period="Отчёт за июль 2026",
    dates=(date(2026, 7, 30), date(2026, 7, 31)),
    duplicate_header=False,
    include_base_rows=True,
    formula_without_cache=False,
):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    sheet.cell(1, 1, period)
    sheet.cell(3, 1, "Дата:")
    sheet.cell(4, 1, "Смена:")

    column = 2
    for report_date in dates:
        sheet.cell(3, column, report_date)
        sheet.cell(4, column, 1)
        sheet.cell(4, column + 1, 2)
        column += 2

    if duplicate_header:
        sheet.cell(3, 4, dates[0])
        sheet.cell(4, 4, 1)

    if include_base_rows:
        sheet.cell(5, 1, " Конвейер\u00a04 ")
        sheet.cell(6, 1, "Конвейер 3")
        for offset in range(len(dates) * 2):
            sheet.cell(5, 2 + offset, "1 000,5")
            sheet.cell(6, 2 + offset, 900)
        if formula_without_cache:
            sheet.cell(5, 2, "=500+500")

    buffer = io.BytesIO()
    workbook.save(buffer)
    workbook.close()
    return buffer.getvalue()


class ExcelParserTests(unittest.TestCase):
    def test_real_xls_binary_is_parsed(self):
        result = main.parse_report(XLS_FIXTURE, "report.xls")

        self.assertEqual(result["error"], "")
        self.assertEqual(result["daily_by_shift"][30][1]["kv4"], 500)
        self.assertEqual(result["daily_by_shift"][31][2]["kv3"], 450)

    def test_xlsx_normalizes_sheet_and_conveyor_labels(self):
        result = main.parse_report(make_xlsx(), "report.xlsx")

        self.assertEqual(result["error"], "")
        self.assertEqual(result["used_sheet"], " ДетСменТекМесяц ")
        self.assertEqual((result["report_year"], result["report_month"]), (2026, 7))
        self.assertEqual(result["daily_by_shift"][30][1]["kv4"], 1000.5)
        self.assertEqual(result["daily_by_shift"][31][2]["kv3"], 900)
        self.assertTrue(
            any("дублирующих весов" in warning for warning in result["warnings"])
        )
        self.assertTrue(
            any("склада концентрата" in warning for warning in result["warnings"])
        )

    def test_uncached_xlsx_formula_is_not_silently_converted_to_zero(self):
        result = main.parse_report(
            make_xlsx(formula_without_cache=True),
            "formula.xlsx",
        )

        self.assertIn("формула не имеет сохранённого", result["error"])
        self.assertEqual(result["daily_by_shift"], {})

    def test_duplicate_day_and_shift_columns_are_rejected(self):
        result = main.parse_report(
            make_xlsx(duplicate_header=True),
            "duplicate.xlsx",
        )

        self.assertIn("повторяются в нескольких колонках", result["error"])

    def test_multiple_months_in_date_row_are_rejected(self):
        result = main.parse_report(
            make_xlsx(dates=(date(2026, 7, 31), date(2026, 8, 1))),
            "mixed.xlsx",
        )

        self.assertIn("несколько месяцев", result["error"])

    def test_missing_expected_sheet_is_reported(self):
        result = main.parse_report(
            make_xlsx(sheet_name="Другой лист"),
            "wrong-sheet.xlsx",
        )

        self.assertIn("не найден лист", result["error"])
        self.assertIn("Другой лист", result["error"])

    def test_missing_input_weight_rows_are_rejected(self):
        result = main.parse_report(
            make_xlsx(include_base_rows=False),
            "missing-base.xlsx",
        )

        self.assertIn("обязательные строки входных весов", result["error"])

    def test_corrupt_workbook_returns_user_facing_error(self):
        result = main.parse_report(b"not an Excel file", "broken.xlsx")

        self.assertIn("Не удалось открыть файл", result["error"])

    def test_invalid_date_or_shift_header_is_rejected(self):
        for row, value, expected in (
            (3, "день один", "значение даты"),
            (4, 3, "номер смены"),
            (4, 1.5, "номер смены"),
        ):
            with self.subTest(row=row, value=value):
                workbook = openpyxl.load_workbook(io.BytesIO(make_xlsx()))
                workbook.active.cell(row, 2, value)
                buffer = io.BytesIO()
                workbook.save(buffer)
                workbook.close()

                result = main.parse_report(buffer.getvalue(), "header.xlsx")
                self.assertIn(expected, result["error"])

    def test_oversized_unpacked_xlsx_is_rejected_before_xml_parsing(self):
        with patch.object(main, "MAX_XLSX_UNCOMPRESSED_BYTES", 1):
            result = main.parse_report(make_xlsx(), "oversized.xlsx")

        self.assertIn("распакованный XLSX", result["error"])


if __name__ == "__main__":
    unittest.main()
